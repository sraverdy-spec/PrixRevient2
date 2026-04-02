from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from io import BytesIO, StringIO
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import asyncio
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"

def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

# ================= AUTH HELPERS =================

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=60), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Non authentifie")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouve")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expire")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

VALID_ROLES = ["admin", "manager", "operator"]

def require_role(*roles):
    """Dependency factory: returns a dependency that checks the user has one of the given roles."""
    async def role_checker(request: Request):
        user = await get_current_user(request)
        if user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="Acces interdit - role insuffisant")
        return user
    return role_checker

require_admin = require_role("admin")
require_manager_or_admin = require_role("admin", "manager")

# ================= MODELS =================

# User Models
class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserCreate(BaseModel):
    """Admin creates a user"""
    email: EmailStr
    password: str
    name: str
    role: str = "operator"

class UserUpdate(BaseModel):
    """Admin updates a user"""
    name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class UserChangePassword(BaseModel):
    new_password: str

# Category Model
class CategoryBase(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#002FA7"

class Category(CategoryBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Supplier Model
class SupplierBase(BaseModel):
    name: str
    code: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None

class Supplier(SupplierBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Raw Material with freinte (waste/loss)
class RawMaterialBase(BaseModel):
    name: str
    code_article: Optional[str] = None
    unit: str
    unit_price: float
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    category_id: Optional[str] = None
    description: Optional[str] = None
    freinte: float = 0.0  # Waste percentage (0-100)
    stock_quantity: float = 0.0
    stock_alert_threshold: float = 0.0

class RawMaterial(RawMaterialBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Recipe Ingredient with sub-recipe support
class RecipeIngredient(BaseModel):
    material_id: Optional[str] = None  # Raw material ID
    sub_recipe_id: Optional[str] = None  # OR Sub-recipe ID for BOM
    material_name: str
    quantity: float
    unit: str
    unit_price: float
    freinte: float = 0.0  # Inherited or overridden
    is_sub_recipe: bool = False

class LaborCost(BaseModel):
    description: str
    hours: float
    hourly_rate: float

class OverheadCostBase(BaseModel):
    name: str
    category: str
    monthly_amount: float
    allocation_method: str
    allocation_value: Optional[float] = 1.0

class OverheadCost(OverheadCostBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Recipe with category and margin
class RecipeBase(BaseModel):
    name: str
    description: Optional[str] = None
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    product_type: Optional[str] = None  # MDD, MN, SM, MP
    version: int = 1
    output_quantity: float = 1.0
    output_unit: str = "pièce"
    ingredients: List[RecipeIngredient] = []
    labor_costs: List[LaborCost] = []
    overhead_ids: List[str] = []
    target_margin: float = 30.0
    is_intermediate: bool = False

class RecipeCreate(RecipeBase):
    pass

class RecipeUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None
    product_type: Optional[str] = None
    version: Optional[int] = None
    output_quantity: Optional[float] = None
    output_unit: Optional[str] = None
    ingredients: Optional[List[RecipeIngredient]] = None
    labor_costs: Optional[List[LaborCost]] = None
    overhead_ids: Optional[List[str]] = None
    target_margin: Optional[float] = None
    is_intermediate: Optional[bool] = None

class Recipe(RecipeBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Price History
class PriceHistory(BaseModel):
    recipe_id: str
    recipe_name: str
    cost_per_unit: float
    total_material_cost: float
    total_labor_cost: float
    total_overhead_cost: float
    total_cost: float
    recorded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Cost Breakdown Response
class CostBreakdown(BaseModel):
    recipe_id: str
    recipe_name: str
    category_id: Optional[str] = None
    supplier_name: Optional[str] = None
    product_type: Optional[str] = None
    total_material_cost: float
    total_labor_cost: float
    total_overhead_cost: float
    total_freinte_cost: float
    total_cost: float
    cost_per_unit: float
    output_quantity: float
    output_unit: str
    target_margin: float
    suggested_price: float
    material_details: List[dict]
    labor_details: List[dict]
    overhead_details: List[dict]
    sub_recipe_details: List[dict] = []

# Simulation
class SimulationRequest(BaseModel):
    recipe_id: str
    material_changes: Dict[str, float] = {}  # material_id: new_price
    labor_rate_change: Optional[float] = None  # percentage change

# ================= AUTH ENDPOINTS =================

@api_router.post("/auth/register")
async def register(input_data: UserRegister, response: Response):
    """Registration disabled - admin creates accounts via /api/users"""
    raise HTTPException(status_code=403, detail="Inscription desactivee. Contactez l'administrateur pour obtenir un compte.")

@api_router.post("/auth/login")
async def login(input_data: UserLogin, request: Request, response: Response):
    email = input_data.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(input_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    if user.get("is_active") is False:
        raise HTTPException(status_code=403, detail="Compte desactive. Contactez l'administrateur.")
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    is_secure = os.environ.get("FRONTEND_URL", "").startswith("https")
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=is_secure, samesite="lax", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=is_secure, samesite="lax", max_age=604800, path="/")
    return {"_id": user_id, "email": user["email"], "name": user["name"], "role": user.get("role", "operator")}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Déconnexion réussie"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    return await get_current_user(request)

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Token de rafraichissement manquant")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Type de token invalide")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilisateur non trouve")
        access_token = create_access_token(str(user["_id"]), user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=is_secure, samesite="lax", max_age=3600, path="/")
        return {"message": "Token rafraichi"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expire")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

# ================= USER MANAGEMENT (Admin Only) =================

@api_router.get("/users")
async def list_users(admin: dict = Depends(require_admin)):
    """List all users (admin only)"""
    users = await db.users.find({}, {"password_hash": 0}).to_list(1000)
    for u in users:
        u["_id"] = str(u["_id"])
    return users

@api_router.post("/users")
async def create_user(input_data: UserCreate, admin: dict = Depends(require_admin)):
    """Create a new user (admin only)"""
    email = input_data.email.lower()
    if input_data.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Role invalide. Roles valides: {', '.join(VALID_ROLES)}")
    
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email deja utilise")
    
    password_hash = hash_password(input_data.password)
    user_doc = {
        "email": email,
        "password_hash": password_hash,
        "name": input_data.name,
        "role": input_data.role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc),
        "created_by": admin["_id"]
    }
    result = await db.users.insert_one(user_doc)
    return {"_id": str(result.inserted_id), "email": email, "name": input_data.name, "role": input_data.role, "is_active": True}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, input_data: UserUpdate, admin: dict = Depends(require_admin)):
    """Update user role or status (admin only)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouve")
    
    # Prevent admin from demoting themselves
    if str(user["_id"]) == admin["_id"] and input_data.role and input_data.role != "admin":
        raise HTTPException(status_code=400, detail="Impossible de modifier votre propre role admin")
    
    update = {}
    if input_data.name is not None:
        update["name"] = input_data.name
    if input_data.role is not None:
        if input_data.role not in VALID_ROLES:
            raise HTTPException(status_code=400, detail=f"Role invalide. Roles valides: {', '.join(VALID_ROLES)}")
        update["role"] = input_data.role
    if input_data.is_active is not None:
        if str(user["_id"]) == admin["_id"]:
            raise HTTPException(status_code=400, detail="Impossible de desactiver votre propre compte")
        update["is_active"] = input_data.is_active
    
    if update:
        update["updated_at"] = datetime.now(timezone.utc)
        await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update})
    
    updated = await db.users.find_one({"_id": ObjectId(user_id)}, {"password_hash": 0})
    updated["_id"] = str(updated["_id"])
    return updated

@api_router.put("/users/{user_id}/password")
async def change_user_password(user_id: str, input_data: UserChangePassword, admin: dict = Depends(require_admin)):
    """Reset a user's password (admin only)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouve")
    
    new_hash = hash_password(input_data.new_password)
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Mot de passe modifie"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    """Delete a user (admin only)"""
    user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouve")
    if str(user["_id"]) == admin["_id"]:
        raise HTTPException(status_code=400, detail="Impossible de supprimer votre propre compte")
    
    await db.users.delete_one({"_id": ObjectId(user_id)})
    return {"message": "Utilisateur supprime"}

# ================= CATEGORIES ENDPOINTS =================

@api_router.get("/categories")
async def get_categories():
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)
    return categories

@api_router.post("/categories")
async def create_category(input_data: CategoryBase):
    category = Category(**input_data.model_dump())
    doc = category.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.categories.insert_one(doc)
    return category

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str):
    await db.categories.delete_one({"id": category_id})
    return {"message": "Catégorie supprimée"}

# ================= SUPPLIERS ENDPOINTS =================

@api_router.get("/suppliers")
async def get_suppliers():
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(input_data: SupplierBase):
    supplier = Supplier(**input_data.model_dump())
    doc = supplier.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.suppliers.insert_one(doc)
    return supplier

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, input_data: SupplierBase):
    await db.suppliers.update_one({"id": supplier_id}, {"$set": input_data.model_dump()})
    return await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str):
    await db.suppliers.delete_one({"id": supplier_id})
    return {"message": "Fournisseur supprimé"}

@api_router.post("/suppliers/import-csv")
async def import_suppliers_csv(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format CSV")
    content = await file.read()
    decoded = content.decode('utf-8-sig')
    reader = csv.DictReader(StringIO(decoded), delimiter=';')
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(StringIO(decoded), delimiter=',')
    imported_count = 0
    errors = []
    for row_num, row in enumerate(reader, start=2):
        try:
            row = {k.strip().lower().replace('\ufeff', ''): v.strip() if v else '' for k, v in row.items() if k}
            name = row.get('name', row.get('nom', '')).strip()
            if not name:
                continue
            supplier = Supplier(
                name=name,
                contact=row.get('contact', ''),
                email=row.get('email', ''),
                phone=row.get('phone', row.get('telephone', '')),
                address=row.get('address', row.get('adresse', '')),
            )
            doc = supplier.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.suppliers.insert_one(doc)
            imported_count += 1
        except Exception as e:
            errors.append(f"Ligne {row_num}: {str(e)}")
    return {"success": imported_count > 0, "imported_count": imported_count, "errors": errors}

@api_router.get("/suppliers/csv-template")
async def get_suppliers_csv_template():
    template = """name;contact;email;phone;address
Moulin du Lac;Jean Dupont;contact@moulin.fr;0145678900;12 rue du Moulin 75001 Paris
Laiterie Centrale;Marie Martin;info@laiterie.fr;0234567890;5 avenue du Lait 69001 Lyon
"""
    return StreamingResponse(
        BytesIO(template.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_fournisseurs.csv"}
    )

@api_router.post("/categories/import-csv")
async def import_categories_csv(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format CSV")
    content = await file.read()
    decoded = content.decode('utf-8-sig')
    reader = csv.DictReader(StringIO(decoded), delimiter=';')
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(StringIO(decoded), delimiter=',')
    imported_count = 0
    errors = []
    for row_num, row in enumerate(reader, start=2):
        try:
            row = {k.strip().lower().replace('\ufeff', ''): v.strip() if v else '' for k, v in row.items() if k}
            name = row.get('name', row.get('nom', '')).strip()
            if not name:
                continue
            color = row.get('color', row.get('couleur', '#002FA7')).strip() or '#002FA7'
            description = row.get('description', '').strip()
            category = Category(name=name, description=description, color=color)
            doc = category.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.categories.insert_one(doc)
            imported_count += 1
        except Exception as e:
            errors.append(f"Ligne {row_num}: {str(e)}")
    return {"success": imported_count > 0, "imported_count": imported_count, "errors": errors}

@api_router.get("/categories/csv-template")
async def get_categories_csv_template():
    template = """name;description;color
Boulangerie;Produits de boulangerie;#002FA7
Patisserie;Gateaux et desserts;#E91E63
Viennoiserie;Croissants et pains au chocolat;#FF9800
"""
    return StreamingResponse(
        BytesIO(template.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_categories.csv"}
    )

# ================= RAW MATERIALS ENDPOINTS =================

@api_router.get("/materials")
async def get_materials():
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(1000)
    for mat in materials:
        if isinstance(mat.get('created_at'), str):
            mat['created_at'] = datetime.fromisoformat(mat['created_at'])
        if isinstance(mat.get('updated_at'), str):
            mat['updated_at'] = datetime.fromisoformat(mat['updated_at'])
    return materials

@api_router.post("/materials")
async def create_material(input_data: RawMaterialBase):
    material = RawMaterial(**input_data.model_dump())
    doc = material.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.raw_materials.insert_one(doc)
    return material

@api_router.put("/materials/{material_id}")
async def update_material(material_id: str, input_data: RawMaterialBase):
    existing = await db.raw_materials.find_one({"id": material_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Matière première non trouvée")
    
    # Track price history if price changed
    old_price = existing.get('unit_price', 0)
    new_price = input_data.unit_price
    if old_price != new_price:
        await db.material_price_history.insert_one({
            "material_id": material_id,
            "old_price": old_price,
            "new_price": new_price,
            "changed_at": datetime.now(timezone.utc).isoformat()
        })
    
    updated_data = input_data.model_dump()
    updated_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.raw_materials.update_one({"id": material_id}, {"$set": updated_data})
    return await db.raw_materials.find_one({"id": material_id}, {"_id": 0})

@api_router.delete("/materials/{material_id}")
async def delete_material(material_id: str):
    await db.raw_materials.delete_one({"id": material_id})
    return {"message": "Matière première supprimée"}

# Import CSV for materials
@api_router.post("/materials/import-csv")
async def import_materials_csv(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format CSV")
    
    content = await file.read()
    decoded = content.decode('utf-8-sig')
    reader = csv.DictReader(StringIO(decoded), delimiter=';')
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(StringIO(decoded), delimiter=',')
    
    imported_count = 0
    errors = []
    
    for row_num, row in enumerate(reader, start=2):
        try:
            row = {k.strip().lower().replace('\ufeff', ''): v.strip() if v else '' for k, v in row.items() if k}
            name = row.get('name', row.get('nom', '')).strip()
            if not name:
                continue
            
            unit = row.get('unit', row.get('unite', 'unité')).strip() or 'unité'
            unit_price = float(row.get('unit_price', row.get('prix_unitaire', row.get('prix', '0'))).replace(',', '.') or 0)
            supplier = row.get('supplier', row.get('fournisseur', '')).strip()
            freinte = float(row.get('freinte', row.get('perte', '0')).replace(',', '.') or 0)
            stock = float(row.get('stock', row.get('stock_quantity', '0')).replace(',', '.') or 0)
            
            material = RawMaterial(
                name=name, unit=unit, unit_price=unit_price,
                supplier_name=supplier, freinte=freinte, stock_quantity=stock
            )
            doc = material.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            doc['updated_at'] = doc['updated_at'].isoformat()
            await db.raw_materials.insert_one(doc)
            imported_count += 1
        except Exception as e:
            errors.append(f"Ligne {row_num}: {str(e)}")
    
    return {"success": imported_count > 0, "imported_count": imported_count, "errors": errors}

@api_router.get("/materials/csv-template")
async def get_materials_csv_template():
    template = """name;unit;unit_price;supplier;freinte;stock
Farine de blé T55;kg;1.20;Moulin du Lac;2;50
Beurre AOP;kg;8.50;Laiterie Centrale;0;20
Sucre en poudre;kg;1.50;Sucre SA;1;30
"""
    return StreamingResponse(
        BytesIO(template.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_matieres.csv"}
    )

# Stock alerts
@api_router.get("/materials/stock-alerts")
async def get_stock_alerts():
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(1000)
    alerts = [m for m in materials if m.get('stock_quantity', 0) <= m.get('stock_alert_threshold', 0) and m.get('stock_alert_threshold', 0) > 0]
    return alerts

# ================= OVERHEAD COSTS ENDPOINTS =================

@api_router.get("/overheads")
async def get_overheads():
    overheads = await db.overheads.find({}, {"_id": 0}).to_list(1000)
    return overheads

@api_router.post("/overheads")
async def create_overhead(input_data: OverheadCostBase):
    overhead = OverheadCost(**input_data.model_dump())
    doc = overhead.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.overheads.insert_one(doc)
    return overhead

@api_router.put("/overheads/{overhead_id}")
async def update_overhead(overhead_id: str, input_data: OverheadCostBase):
    await db.overheads.update_one({"id": overhead_id}, {"$set": input_data.model_dump()})
    return await db.overheads.find_one({"id": overhead_id}, {"_id": 0})

@api_router.delete("/overheads/{overhead_id}")
async def delete_overhead(overhead_id: str):
    await db.overheads.delete_one({"id": overhead_id})
    return {"message": "Frais général supprimé"}

# ================= RECIPES ENDPOINTS =================

@api_router.get("/recipes")
async def get_recipes():
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
    return recipes

@api_router.get("/recipes/intermediate")
async def get_intermediate_recipes():
    """Get recipes that can be used as sub-recipes (BOM)"""
    recipes = await db.recipes.find({"is_intermediate": True}, {"_id": 0}).to_list(1000)
    return recipes

@api_router.get("/recipes/csv-template")
async def get_recipes_csv_template():
    template = """name;description;output_quantity;output_unit;margin;is_intermediate;ingredient_name;ingredient_quantity;ingredient_unit;ingredient_price;freinte;sub_recipe;labor_description;labor_hours;labor_rate
Pate brisee;Pate de base;1;kg;30;oui;Farine de ble;0.5;kg;1.20;2;;Petrissage;0.5;15
Pate brisee;Pate de base;1;kg;30;oui;Beurre;0.25;kg;8.00;3;;;;
Tarte aux pommes;Tarte classique;8;piece;35;non;Sucre;0.1;kg;1.50;2;;Preparation;1;15
Tarte aux pommes;Tarte classique;8;piece;35;non;Oeufs;3;piece;0.25;0;;Cuisson;0.75;12
Tarte aux pommes;Tarte classique;8;piece;35;non;;;;;;Pate brisee:0.5:kg;;;
"""
    return StreamingResponse(
        BytesIO(template.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_recettes_arbre.csv"}
    )

@api_router.get("/recipes/bom-csv-template")
async def get_bom_csv_template():
    template = """name;description;output_quantity;output_unit;margin;is_intermediate;ingredient_name;ingredient_quantity;ingredient_unit;ingredient_price;freinte;sub_recipe;labor_description;labor_hours;labor_rate
Pate brisee;Pate de base;1;kg;30;oui;Farine de ble;0.5;kg;1.20;2;;Petrissage;0.5;15
Pate brisee;Pate de base;1;kg;30;oui;Beurre;0.25;kg;8.00;3;;;;
Tarte aux pommes;Tarte classique;8;piece;35;non;Sucre;0.1;kg;1.50;2;;Preparation;1;15
Tarte aux pommes;Tarte classique;8;piece;35;non;Oeufs;3;piece;0.25;0;;Cuisson;0.75;12
Tarte aux pommes;Tarte classique;8;piece;35;non;;;;;;Pate brisee:0.5:kg;;;
"""
    return StreamingResponse(
        BytesIO(template.encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_arbre_fabrication.csv"}
    )

@api_router.get("/recipes/{recipe_id}")
async def get_recipe(recipe_id: str):
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recette non trouvée")
    return recipe

@api_router.post("/recipes")
async def create_recipe(input_data: RecipeCreate):
    recipe = Recipe(**input_data.model_dump())
    doc = recipe.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.recipes.insert_one(doc)
    return recipe

@api_router.put("/recipes/{recipe_id}")
async def update_recipe(recipe_id: str, input_data: RecipeUpdate):
    existing = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Recette non trouvée")
    
    update_data = {k: v for k, v in input_data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.recipes.update_one({"id": recipe_id}, {"$set": update_data})
    return await db.recipes.find_one({"id": recipe_id}, {"_id": 0})

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str):
    await db.recipes.delete_one({"id": recipe_id})
    return {"message": "Recette supprimée"}

@api_router.post("/recipes/{recipe_id}/duplicate")
async def duplicate_recipe(recipe_id: str):
    """Duplicate a recipe with incremented version"""
    original = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Recette non trouvée")
    
    base_name = original["name"]
    # Find max version for recipes with the same name
    same_name = await db.recipes.find({"name": base_name}, {"_id": 0, "version": 1}).to_list(100)
    max_version = max((r.get("version", 1) for r in same_name), default=1)
    
    new_recipe = {**original}
    new_recipe["id"] = str(uuid.uuid4())
    new_recipe["version"] = max_version + 1
    new_recipe["created_at"] = datetime.now(timezone.utc).isoformat()
    new_recipe["updated_at"] = datetime.now(timezone.utc).isoformat()
    new_recipe.pop("_id", None)
    
    await db.recipes.insert_one(new_recipe)
    result = await db.recipes.find_one({"id": new_recipe["id"]}, {"_id": 0})
    return result

# ================= COST CALCULATION =================

async def calculate_sub_recipe_cost(recipe_id: str, depth: int = 0) -> dict:
    """Recursively calculate cost for sub-recipes (BOM)"""
    if depth > 10:  # Prevent infinite recursion
        return {"cost": 0, "details": [], "ingredients": []}
    
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        return {"cost": 0, "details": [], "ingredients": []}
    
    total_cost = 0
    material_cost = 0
    labor_cost = 0
    freinte_cost = 0
    ingredients_detail = []
    
    # Calculate material costs including freinte
    for ing in recipe.get('ingredients', []):
        if ing.get('is_sub_recipe') and ing.get('sub_recipe_id'):
            # Recursive call for sub-recipe
            sub_cost = await calculate_sub_recipe_cost(ing['sub_recipe_id'], depth + 1)
            cost = ing['quantity'] * sub_cost['cost']
            ingredients_detail.append({
                "name": ing['material_name'], "quantity": ing['quantity'], "unit": ing.get('unit', ''),
                "is_sub_recipe": True, "total_cost": round(cost, 2),
                "sub_ingredients": sub_cost.get('ingredients', []),
            })
        else:
            base_cost = ing['quantity'] * ing['unit_price']
            freinte_pct = ing.get('freinte', 0) / 100
            freinte_add = base_cost * freinte_pct
            cost = base_cost + freinte_add
            freinte_cost += freinte_add
            # Lookup code_article for sub-recipe material
            sub_mat_doc = await db.raw_materials.find_one({"id": ing.get('material_id')}, {"_id": 0, "code_article": 1})
            ingredients_detail.append({
                "name": ing['material_name'], "code_article": sub_mat_doc.get('code_article', '') if sub_mat_doc else '',
                "quantity": ing['quantity'], "unit": ing.get('unit', ''),
                "unit_price": ing['unit_price'], "freinte": ing.get('freinte', 0),
                "is_sub_recipe": False, "total_cost": round(cost, 2),
            })
        material_cost += cost
    
    # Calculate labor costs
    for labor in recipe.get('labor_costs', []):
        labor_cost += labor['hours'] * labor['hourly_rate']
    
    total_cost = material_cost + labor_cost
    output_qty = recipe.get('output_quantity', 1) or 1
    
    return {
        "cost": total_cost / output_qty,
        "material_cost": material_cost,
        "labor_cost": labor_cost,
        "freinte_cost": freinte_cost,
        "ingredients": ingredients_detail,
    }

async def calculate_cost(recipe_id: str) -> CostBreakdown:
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recette non trouvée")
    
    # Calculate material costs with freinte
    material_details = []
    sub_recipe_details = []
    total_material_cost = 0.0
    total_freinte_cost = 0.0
    
    for ing in recipe.get('ingredients', []):
        if ing.get('is_sub_recipe') and ing.get('sub_recipe_id'):
            # Sub-recipe cost
            sub_cost_data = await calculate_sub_recipe_cost(ing['sub_recipe_id'])
            cost = ing['quantity'] * sub_cost_data['cost']
            total_material_cost += cost
            sub_recipe_details.append({
                "name": ing['material_name'],
                "quantity": ing['quantity'],
                "unit": ing['unit'],
                "unit_cost": sub_cost_data['cost'],
                "total_cost": round(cost, 2),
                "material_cost": round(sub_cost_data.get('material_cost', 0), 2),
                "labor_cost": round(sub_cost_data.get('labor_cost', 0), 2),
                "ingredients": sub_cost_data.get('ingredients', []),
            })
        else:
            base_cost = ing['quantity'] * ing['unit_price']
            freinte_pct = ing.get('freinte', 0) / 100
            freinte_add = base_cost * freinte_pct
            cost = base_cost + freinte_add
            total_material_cost += cost
            total_freinte_cost += freinte_add
            # Lookup code_article from raw_materials
            mat_doc = await db.raw_materials.find_one({"id": ing.get('material_id')}, {"_id": 0, "code_article": 1})
            material_details.append({
                "name": ing['material_name'],
                "code_article": mat_doc.get('code_article', '') if mat_doc else '',
                "quantity": ing['quantity'],
                "unit": ing['unit'],
                "unit_price": ing['unit_price'],
                "freinte": ing.get('freinte', 0),
                "freinte_cost": round(freinte_add, 2),
                "total_cost": round(cost, 2)
            })
    
    # Calculate labor costs
    labor_details = []
    total_labor_cost = 0.0
    for labor in recipe.get('labor_costs', []):
        cost = labor['hours'] * labor['hourly_rate']
        total_labor_cost += cost
        labor_details.append({
            "description": labor['description'],
            "hours": labor['hours'],
            "hourly_rate": labor['hourly_rate'],
            "total_cost": round(cost, 2)
        })
    
    # Calculate overhead costs
    overhead_details = []
    total_overhead_cost = 0.0
    total_labor_hours = sum(l['hours'] for l in recipe.get('labor_costs', []))
    
    for oh_id in recipe.get('overhead_ids', []):
        overhead = await db.overheads.find_one({"id": oh_id}, {"_id": 0})
        if overhead:
            if overhead['allocation_method'] == 'per_unit':
                cost = overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)
            elif overhead['allocation_method'] == 'per_hour':
                cost = (overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)) * total_labor_hours
            else:
                cost = overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)
            
            total_overhead_cost += cost
            overhead_details.append({
                "name": overhead['name'],
                "category": overhead['category'],
                "allocation_method": overhead['allocation_method'],
                "total_cost": round(cost, 2)
            })
    
    total_cost = total_material_cost + total_labor_cost + total_overhead_cost
    output_qty = recipe.get('output_quantity', 1) or 1
    cost_per_unit = total_cost / output_qty
    target_margin = recipe.get('target_margin', 30)
    suggested_price = cost_per_unit / (1 - target_margin / 100) if target_margin < 100 else cost_per_unit * 2
    
    return CostBreakdown(
        recipe_id=recipe_id,
        recipe_name=recipe['name'],
        category_id=recipe.get('category_id'),
        supplier_name=recipe.get('supplier_name', ''),
        product_type=recipe.get('product_type', ''),
        total_material_cost=round(total_material_cost, 2),
        total_labor_cost=round(total_labor_cost, 2),
        total_overhead_cost=round(total_overhead_cost, 2),
        total_freinte_cost=round(total_freinte_cost, 2),
        total_cost=round(total_cost, 2),
        cost_per_unit=round(cost_per_unit, 2),
        output_quantity=output_qty,
        output_unit=recipe.get('output_unit', 'pièce'),
        target_margin=target_margin,
        suggested_price=round(suggested_price, 2),
        material_details=material_details,
        labor_details=labor_details,
        overhead_details=overhead_details,
        sub_recipe_details=sub_recipe_details
    )

@api_router.get("/recipes/{recipe_id}/cost")
async def get_recipe_cost(recipe_id: str):
    return await calculate_cost(recipe_id)

# ========= SIMULATION VERSIONS =========
@api_router.get("/recipes/{recipe_id}/simulations")
async def get_simulations(recipe_id: str):
    """Get all simulation versions for a recipe"""
    sims = await db.recipe_simulations.find(
        {"recipe_id": recipe_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return sims

@api_router.post("/recipes/{recipe_id}/simulations")
async def save_simulation(recipe_id: str, request: Request, user: dict = Depends(require_role("admin", "manager"))):
    """Save a simulation version"""
    body = await request.json()
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recette introuvable")

    # Count existing versions
    count = await db.recipe_simulations.count_documents({"recipe_id": recipe_id})
    version_num = count + 1

    sim = {
        "id": str(uuid.uuid4()),
        "recipe_id": recipe_id,
        "version": version_num,
        "label": body.get("label", f"Simulation v{version_num}"),
        "sim_ingredients": body.get("sim_ingredients", {}),
        "sim_labor": body.get("sim_labor", {}),
        "cost_summary": body.get("cost_summary", {}),
        "created_by": user.get("email", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.recipe_simulations.insert_one(sim)
    # Remove _id
    return {"message": "Simulation sauvegardee", "simulation": {k: v for k, v in sim.items() if k != "_id"}}

@api_router.delete("/recipes/{recipe_id}/simulations/{sim_id}")
async def delete_simulation(recipe_id: str, sim_id: str, user: dict = Depends(require_admin)):
    await db.recipe_simulations.delete_one({"id": sim_id, "recipe_id": recipe_id})
    return {"message": "Simulation supprimee"}


# ========= RECIPE IMAGE UPLOAD =========
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

@api_router.post("/recipes/{recipe_id}/image")
async def upload_recipe_image(recipe_id: str, file: UploadFile = File(...)):
    """Upload an image for a recipe"""
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recette introuvable")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "jpg"
    if ext not in ("jpg", "jpeg", "png", "webp"):
        raise HTTPException(status_code=400, detail="Format image non supporte (jpg, png, webp)")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image trop volumineuse (max 5 Mo)")

    filename = f"{recipe_id}.{ext}"
    filepath = UPLOAD_DIR / filename
    filepath.write_bytes(content)

    image_url = f"/api/uploads/{filename}"
    await db.recipes.update_one({"id": recipe_id}, {"$set": {"image_url": image_url}})
    return {"image_url": image_url}

@api_router.post("/recipes/{recipe_id}/image-url")
async def set_recipe_image_url(recipe_id: str, request: Request):
    """Set an image URL for a recipe"""
    body = await request.json()
    url = body.get("url", "")
    if not url:
        raise HTTPException(status_code=400, detail="URL requise")
    await db.recipes.update_one({"id": recipe_id}, {"$set": {"image_url": url}})
    return {"image_url": url}

@api_router.get("/uploads/{filename}")
async def serve_upload(filename: str):
    """Serve uploaded files"""
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404)
    ext = filename.split(".")[-1].lower()
    media_types = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
    return StreamingResponse(open(filepath, "rb"), media_type=media_types.get(ext, "application/octet-stream"))

# ========= INDIVIDUAL RECIPE EXCEL EXPORT =========
@api_router.get("/recipes/{recipe_id}/excel")
async def export_recipe_excel(recipe_id: str):
    """Export a single recipe as Excel"""
    cost = await calculate_cost(recipe_id)
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})

    wb = Workbook()
    ws = wb.active
    ws.title = "Fiche recette"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    amber_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Header info
    ws.merge_cells("A1:E1")
    ws["A1"] = f"FICHE DE PRIX DE REVIENT - {cost.recipe_name}"
    ws["A1"].font = Font(bold=True, size=16, color="002FA7")
    ws["A2"] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A3"] = f"Production: {cost.output_quantity} {cost.output_unit}"
    ws["A4"] = f"Client: {recipe.get('supplier_name', '-')}"
    ws["A5"] = f"Type: {recipe.get('product_type', '-')}"

    row = 7
    # Materials
    ws.cell(row=row, column=1, value="MATIERES PREMIERES").font = bold_font
    row += 1
    for col, h in enumerate(["Code article", "Matiere", "Quantite", "Unite", "Prix/u", "Freinte %", "Total"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
    row += 1
    for m in cost.material_details:
        ws.cell(row=row, column=1, value=m.get("code_article", "")).border = border
        ws.cell(row=row, column=2, value=m["name"]).border = border
        ws.cell(row=row, column=3, value=m["quantity"]).border = border
        ws.cell(row=row, column=4, value=m["unit"]).border = border
        ws.cell(row=row, column=5, value=m["unit_price"]).border = border
        ws.cell(row=row, column=6, value=m.get("freinte", 0)).border = border
        ws.cell(row=row, column=7, value=m["total_cost"]).border = border
        row += 1
    ws.cell(row=row, column=6, value="Sous-total:").font = bold_font
    ws.cell(row=row, column=7, value=cost.total_material_cost).font = bold_font
    row += 2

    # Sub-recipes
    if cost.sub_recipe_details:
        ws.cell(row=row, column=1, value="ARTICLES SEMI-FINIS").font = bold_font
        row += 1
        for col, h in enumerate(["Semi-fini", "Quantite", "Unite", "Cout/u", "Total"], 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = amber_fill
            c.border = border
        row += 1
        for s in cost.sub_recipe_details:
            ws.cell(row=row, column=1, value=s["name"]).border = border
            ws.cell(row=row, column=2, value=s["quantity"]).border = border
            ws.cell(row=row, column=3, value=s["unit"]).border = border
            ws.cell(row=row, column=4, value=round(s["unit_cost"], 2)).border = border
            ws.cell(row=row, column=5, value=s["total_cost"]).border = border
            row += 1
        row += 1

    # Labor
    ws.cell(row=row, column=1, value="MAIN D'OEUVRE").font = bold_font
    row += 1
    for col, h in enumerate(["Description", "Heures", "Taux horaire", "Total"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = green_fill
        c.border = border
    row += 1
    for l in cost.labor_details:
        ws.cell(row=row, column=1, value=l["description"]).border = border
        ws.cell(row=row, column=2, value=l["hours"]).border = border
        ws.cell(row=row, column=3, value=l["hourly_rate"]).border = border
        ws.cell(row=row, column=4, value=l["total_cost"]).border = border
        row += 1
    row += 1

    # Summary
    ws.cell(row=row, column=1, value="RECAPITULATIF").font = Font(bold=True, size=13, color="002FA7")
    row += 1
    for label, val in [
        ("Cout matieres", cost.total_material_cost),
        ("Cout main d'oeuvre", cost.total_labor_cost),
        ("Frais generaux", cost.total_overhead_cost),
        ("COUT TOTAL", cost.total_cost),
        (f"Prix de revient / {cost.output_unit}", cost.cost_per_unit),
        (f"Marge cible", f"{cost.target_margin}%"),
        ("PRIX DE VENTE CONSEILLE", cost.suggested_price),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=val).font = bold_font
        row += 1

    # Column widths
    for col_letter, width in [("A", 14), ("B", 25), ("C", 10), ("D", 8), ("E", 12), ("F", 10), ("G", 12)]:
        ws.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"fiche_{cost.recipe_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={filename}"})




# Save price history
@api_router.post("/recipes/{recipe_id}/save-history")
async def save_price_history(recipe_id: str):
    cost = await calculate_cost(recipe_id)
    history = PriceHistory(
        recipe_id=recipe_id,
        recipe_name=cost.recipe_name,
        cost_per_unit=cost.cost_per_unit,
        total_material_cost=cost.total_material_cost,
        total_labor_cost=cost.total_labor_cost,
        total_overhead_cost=cost.total_overhead_cost,
        total_cost=cost.total_cost
    )
    doc = history.model_dump()
    doc['recorded_at'] = doc['recorded_at'].isoformat()
    await db.price_history.insert_one(doc)
    return {"message": "Historique enregistré"}

@api_router.get("/recipes/{recipe_id}/history")
async def get_price_history(recipe_id: str):
    history = await db.price_history.find({"recipe_id": recipe_id}, {"_id": 0}).sort("recorded_at", -1).to_list(100)
    return history

# ================= SIMULATION (What-if) =================

@api_router.post("/recipes/{recipe_id}/simulate")
async def simulate_cost(recipe_id: str, simulation: SimulationRequest):
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recette non trouvée")
    
    # Calculate with modified prices
    total_material_cost = 0.0
    for ing in recipe.get('ingredients', []):
        material_id = ing.get('material_id')
        unit_price = simulation.material_changes.get(material_id, ing['unit_price'])
        base_cost = ing['quantity'] * unit_price
        freinte_pct = ing.get('freinte', 0) / 100
        total_material_cost += base_cost * (1 + freinte_pct)
    
    total_labor_cost = 0.0
    rate_multiplier = 1 + (simulation.labor_rate_change or 0) / 100
    for labor in recipe.get('labor_costs', []):
        total_labor_cost += labor['hours'] * labor['hourly_rate'] * rate_multiplier
    
    # Overhead costs (unchanged)
    total_overhead_cost = 0.0
    total_labor_hours = sum(l['hours'] for l in recipe.get('labor_costs', []))
    for oh_id in recipe.get('overhead_ids', []):
        overhead = await db.overheads.find_one({"id": oh_id}, {"_id": 0})
        if overhead:
            if overhead['allocation_method'] == 'per_unit':
                total_overhead_cost += overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)
            elif overhead['allocation_method'] == 'per_hour':
                total_overhead_cost += (overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)) * total_labor_hours
            else:
                total_overhead_cost += overhead['monthly_amount'] / max(overhead.get('allocation_value', 1), 1)
    
    total_cost = total_material_cost + total_labor_cost + total_overhead_cost
    output_qty = recipe.get('output_quantity', 1) or 1
    simulated_cost_per_unit = total_cost / output_qty
    
    # Get original cost for comparison
    original_cost = await calculate_cost(recipe_id)
    
    return {
        "original_cost_per_unit": original_cost.cost_per_unit,
        "simulated_cost_per_unit": round(simulated_cost_per_unit, 2),
        "difference": round(simulated_cost_per_unit - original_cost.cost_per_unit, 2),
        "difference_percentage": round((simulated_cost_per_unit - original_cost.cost_per_unit) / original_cost.cost_per_unit * 100, 2) if original_cost.cost_per_unit > 0 else 0,
        "simulated_material_cost": round(total_material_cost, 2),
        "simulated_labor_cost": round(total_labor_cost, 2),
        "simulated_overhead_cost": round(total_overhead_cost, 2)
    }

# ================= COMPARISON =================

@api_router.post("/recipes/compare")
async def compare_recipes(recipe_ids: List[str]):
    results = []
    for rid in recipe_ids:
        try:
            recipe = await db.recipes.find_one({"id": rid}, {"_id": 0})
            cost = await calculate_cost(rid)
            results.append({
                "recipe_id": rid,
                "recipe_name": cost.recipe_name,
                "supplier_name": recipe.get("supplier_name", "") if recipe else "",
                "version": recipe.get("version", 1) if recipe else 1,
                "cost_per_unit": cost.cost_per_unit,
                "total_cost": cost.total_cost,
                "material_cost": cost.total_material_cost,
                "labor_cost": cost.total_labor_cost,
                "overhead_cost": cost.total_overhead_cost,
                "suggested_price": cost.suggested_price,
                "output_quantity": cost.output_quantity,
                "output_unit": cost.output_unit
            })
        except:
            continue
    return results

# ================= GLOBAL COSTS TABLE =================

@api_router.get("/reports/all-costs")
async def get_all_costs():
    """Get a comprehensive table of all costs for all recipes"""
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
    results = []
    
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe['id'])
            results.append({
                "recipe_id": recipe['id'],
                "recipe_name": recipe['name'],
                "category_id": recipe.get('category_id'),
                "supplier_name": recipe.get('supplier_name', ''),
                "product_type": recipe.get('product_type', ''),
                "version": recipe.get('version', 1),
                "output_quantity": recipe.get('output_quantity', 1),
                "output_unit": recipe.get('output_unit', 'pièce'),
                "material_cost": cost.total_material_cost,
                "labor_cost": cost.total_labor_cost,
                "overhead_cost": cost.total_overhead_cost,
                "freinte_cost": cost.total_freinte_cost,
                "total_cost": cost.total_cost,
                "cost_per_unit": cost.cost_per_unit,
                "target_margin": cost.target_margin,
                "suggested_price": cost.suggested_price
            })
        except Exception as e:
            logging.error(f"Error calculating cost for recipe {recipe['id']}: {e}")
            continue
    
    return results

# ================= CSV IMPORT FOR RECIPES =================

@api_router.post("/recipes/import-csv")
async def import_recipes_csv(file: UploadFile = File(...)):
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format CSV")
    
    content = await file.read()
    decoded = content.decode('utf-8-sig')
    reader = csv.DictReader(StringIO(decoded), delimiter=';')
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(StringIO(decoded), delimiter=',')
    
    recipes_dict = {}
    errors = []
    
    for row_num, row in enumerate(reader, start=2):
        try:
            row = {k.strip().lower().replace('\ufeff', ''): v.strip() if v else '' for k, v in row.items() if k}
            recipe_name = row.get('name', row.get('nom', '')).strip()
            if not recipe_name:
                continue
            
            if recipe_name not in recipes_dict:
                recipes_dict[recipe_name] = {
                    'name': recipe_name,
                    'description': row.get('description', ''),
                    'output_quantity': float(row.get('output_quantity', row.get('quantite_produite', '1')) or 1),
                    'output_unit': row.get('output_unit', row.get('unite_sortie', 'pièce')) or 'pièce',
                    'target_margin': float(row.get('margin', row.get('marge', '30')) or 30),
                    'ingredients': [],
                    'labor_costs': [],
                    'overhead_ids': []
                }
            
            recipe = recipes_dict[recipe_name]
            
            # Add ingredient
            ing_name = row.get('ingredient_name', row.get('matiere', row.get('ingredient', ''))).strip()
            ing_qty = row.get('ingredient_quantity', row.get('quantite', row.get('qte', ''))).strip()
            
            if ing_name and ing_qty:
                ing_unit = row.get('ingredient_unit', row.get('unite', 'unité')).strip() or 'unité'
                ing_price = row.get('ingredient_price', row.get('prix_unitaire', row.get('prix', '0'))).strip()
                freinte = row.get('freinte', row.get('perte', '0')).strip()
                
                # Check if material exists
                existing_mat = await db.raw_materials.find_one({"name": ing_name}, {"_id": 0})
                if not existing_mat:
                    new_mat = RawMaterial(name=ing_name, unit=ing_unit, unit_price=float(ing_price.replace(',', '.')) if ing_price else 0)
                    mat_doc = new_mat.model_dump()
                    mat_doc['created_at'] = mat_doc['created_at'].isoformat()
                    mat_doc['updated_at'] = mat_doc['updated_at'].isoformat()
                    await db.raw_materials.insert_one(mat_doc)
                    existing_mat = mat_doc
                
                recipe['ingredients'].append({
                    'material_id': existing_mat['id'],
                    'material_name': ing_name,
                    'quantity': float(ing_qty.replace(',', '.')),
                    'unit': existing_mat.get('unit', ing_unit),
                    'unit_price': existing_mat.get('unit_price', float(ing_price.replace(',', '.')) if ing_price else 0),
                    'freinte': float(freinte.replace(',', '.')) if freinte else existing_mat.get('freinte', 0),
                    'is_sub_recipe': False
                })
            
            # Add labor
            labor_desc = row.get('labor_description', row.get('travail', row.get('main_oeuvre', ''))).strip()
            labor_hours = row.get('labor_hours', row.get('heures', '')).strip()
            labor_rate = row.get('labor_rate', row.get('taux_horaire', row.get('taux', ''))).strip()
            
            if labor_desc and labor_hours:
                recipe['labor_costs'].append({
                    'description': labor_desc,
                    'hours': float(labor_hours.replace(',', '.')),
                    'hourly_rate': float(labor_rate.replace(',', '.')) if labor_rate else 15.0
                })
                
        except Exception as e:
            errors.append(f"Ligne {row_num}: {str(e)}")
    
    # Save recipes
    imported_recipes = []
    for recipe_data in recipes_dict.values():
        recipe = Recipe(**recipe_data)
        doc = recipe.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.recipes.insert_one(doc)
        imported_recipes.append({'id': recipe.id, 'name': recipe.name})
    
    return {"success": len(imported_recipes) > 0, "imported_count": len(imported_recipes), "errors": errors, "recipes": imported_recipes}

# ================= BOM TREE CSV IMPORT =================

@api_router.post("/recipes/import-bom-csv")
async def import_bom_csv(file: UploadFile = File(...)):
    """Import recipes with manufacturing tree (BOM) from CSV.
    Supports sub_recipe column to link intermediate recipes as ingredients.
    Format sub_recipe: 'RecipeName:quantity:unit'
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit etre au format CSV")
    
    content = await file.read()
    decoded = content.decode('utf-8-sig')
    reader = csv.DictReader(StringIO(decoded), delimiter=';')
    if not reader.fieldnames or len(reader.fieldnames) <= 1:
        reader = csv.DictReader(StringIO(decoded), delimiter=',')
    
    recipes_dict = {}
    sub_recipe_links = {}
    errors = []
    
    rows = list(reader)
    
    # Phase 1: Parse all recipes
    for row_num, row in enumerate(rows, start=2):
        try:
            row = {k.strip().lower().replace('\ufeff', ''): v.strip() if v else '' for k, v in row.items() if k}
            recipe_name = row.get('name', row.get('nom', '')).strip()
            if not recipe_name:
                continue
            
            is_intermediate_str = row.get('is_intermediate', row.get('semi_fini', '')).strip().lower()
            is_intermediate = is_intermediate_str in ('oui', 'true', '1', 'yes', 'vrai')
            
            if recipe_name not in recipes_dict:
                recipes_dict[recipe_name] = {
                    'name': recipe_name,
                    'description': row.get('description', ''),
                    'output_quantity': float(row.get('output_quantity', row.get('quantite_produite', '1')).replace(',', '.') or 1),
                    'output_unit': row.get('output_unit', row.get('unite_sortie', 'piece')) or 'piece',
                    'target_margin': float(row.get('margin', row.get('marge', '30')).replace(',', '.') or 30),
                    'is_intermediate': is_intermediate,
                    'ingredients': [],
                    'labor_costs': [],
                    'overhead_ids': []
                }
                sub_recipe_links[recipe_name] = []
            
            recipe = recipes_dict[recipe_name]
            
            # Check for sub_recipe reference (format: RecipeName:quantity:unit)
            sub_ref = row.get('sub_recipe', row.get('sous_recette', '')).strip()
            if sub_ref:
                parts = sub_ref.split(':')
                sub_name = parts[0].strip()
                sub_qty = float(parts[1].replace(',', '.')) if len(parts) > 1 and parts[1].strip() else 1.0
                sub_unit = parts[2].strip() if len(parts) > 2 and parts[2].strip() else 'unite'
                sub_recipe_links[recipe_name].append({
                    'sub_name': sub_name,
                    'quantity': sub_qty,
                    'unit': sub_unit
                })
            
            # Add raw material ingredient
            ing_name = row.get('ingredient_name', row.get('matiere', row.get('ingredient', ''))).strip()
            ing_qty = row.get('ingredient_quantity', row.get('quantite', row.get('qte', ''))).strip()
            
            if ing_name and ing_qty:
                ing_unit = row.get('ingredient_unit', row.get('unite', 'unite')).strip() or 'unite'
                ing_price = row.get('ingredient_price', row.get('prix_unitaire', row.get('prix', '0'))).strip()
                freinte = row.get('freinte', row.get('perte', '0')).strip()
                
                existing_mat = await db.raw_materials.find_one({"name": ing_name}, {"_id": 0})
                if not existing_mat:
                    new_mat = RawMaterial(
                        name=ing_name, unit=ing_unit,
                        unit_price=float(ing_price.replace(',', '.')) if ing_price else 0,
                        freinte=float(freinte.replace(',', '.')) if freinte else 0
                    )
                    mat_doc = new_mat.model_dump()
                    mat_doc['created_at'] = mat_doc['created_at'].isoformat()
                    mat_doc['updated_at'] = mat_doc['updated_at'].isoformat()
                    await db.raw_materials.insert_one(mat_doc)
                    existing_mat = mat_doc
                
                recipe['ingredients'].append({
                    'material_id': existing_mat['id'],
                    'material_name': ing_name,
                    'quantity': float(ing_qty.replace(',', '.')),
                    'unit': existing_mat.get('unit', ing_unit),
                    'unit_price': existing_mat.get('unit_price', float(ing_price.replace(',', '.')) if ing_price else 0),
                    'freinte': float(freinte.replace(',', '.')) if freinte else existing_mat.get('freinte', 0),
                    'is_sub_recipe': False
                })
            
            # Add labor
            labor_desc = row.get('labor_description', row.get('travail', row.get('main_oeuvre', ''))).strip()
            labor_hours = row.get('labor_hours', row.get('heures', '')).strip()
            labor_rate = row.get('labor_rate', row.get('taux_horaire', row.get('taux', ''))).strip()
            
            if labor_desc and labor_hours:
                recipe['labor_costs'].append({
                    'description': labor_desc,
                    'hours': float(labor_hours.replace(',', '.')),
                    'hourly_rate': float(labor_rate.replace(',', '.')) if labor_rate else 15.0
                })
                
        except Exception as e:
            errors.append(f"Ligne {row_num}: {str(e)}")
    
    # Phase 2: Save intermediate recipes first
    saved_ids = {}
    intermediate_recipes = {k: v for k, v in recipes_dict.items() if v.get('is_intermediate')}
    final_recipes = {k: v for k, v in recipes_dict.items() if not v.get('is_intermediate')}
    
    for recipe_name, recipe_data in intermediate_recipes.items():
        recipe = Recipe(**recipe_data)
        doc = recipe.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.recipes.insert_one(doc)
        saved_ids[recipe_name] = recipe.id
    
    # Phase 3: Save final recipes with sub-recipe references
    for recipe_name, recipe_data in final_recipes.items():
        links = sub_recipe_links.get(recipe_name, [])
        for link in links:
            sub_id = saved_ids.get(link['sub_name'])
            if not sub_id:
                existing_sub = await db.recipes.find_one({"name": link['sub_name'], "is_intermediate": True}, {"_id": 0})
                if existing_sub:
                    sub_id = existing_sub['id']
            
            if sub_id:
                recipe_data['ingredients'].append({
                    'sub_recipe_id': sub_id,
                    'material_name': link['sub_name'],
                    'quantity': link['quantity'],
                    'unit': link['unit'],
                    'unit_price': 0,
                    'freinte': 0,
                    'is_sub_recipe': True
                })
            else:
                errors.append(f"Sous-recette '{link['sub_name']}' introuvable pour '{recipe_name}'")
        
        recipe = Recipe(**recipe_data)
        doc = recipe.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.recipes.insert_one(doc)
        saved_ids[recipe_name] = recipe.id
    
    return {
        "success": len(saved_ids) > 0,
        "imported_count": len(saved_ids),
        "intermediate_count": len(intermediate_recipes),
        "final_count": len(final_recipes),
        "errors": errors,
        "recipes": [{"id": v, "name": k} for k, v in saved_ids.items()]
    }

# ================= AUTO IMPORT API =================

SFTP_WATCH_DIR = os.path.join(ROOT_DIR, "import_watch")
os.makedirs(SFTP_WATCH_DIR, exist_ok=True)
IMPORT_LOG_FILE = os.path.join(ROOT_DIR, "import_log.json")

@api_router.post("/import/auto")
async def auto_import(file: UploadFile = File(...), import_type: str = "materials"):
    """API endpoint for automated CSV imports. 
    Use import_type='materials' or 'recipes' or 'bom'.
    Can be called from external systems (ERP, scripts, SFTP watchers).
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit etre au format CSV")
    
    log_entry = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "filename": file.filename,
        "import_type": import_type,
        "source": "api"
    }
    
    if import_type == "materials":
        result = await import_materials_csv(file)
    elif import_type == "bom":
        result = await import_bom_csv(file)
    elif import_type == "recipes":
        result = await import_recipes_csv(file)
    elif import_type == "suppliers":
        result = await import_suppliers_csv(file)
    elif import_type == "categories":
        result = await import_categories_csv(file)
    else:
        raise HTTPException(status_code=400, detail="Type d'import invalide. Utilisez 'materials', 'recipes', 'bom', 'suppliers' ou 'categories'")
    
    log_entry["result"] = result
    _append_import_log(log_entry)
    
    return result

@api_router.post("/import/sftp-scan")
async def sftp_scan():
    """Scan the SFTP watch directory for CSV files and import them automatically.
    Files are moved to a 'processed' subdirectory after import.
    Naming convention: materials_*.csv, recipes_*.csv, bom_*.csv
    """
    import shutil
    processed_dir = os.path.join(SFTP_WATCH_DIR, "processed")
    os.makedirs(processed_dir, exist_ok=True)
    
    results = []
    files_found = [f for f in os.listdir(SFTP_WATCH_DIR) if f.endswith('.csv')]
    
    for filename in sorted(files_found):
        filepath = os.path.join(SFTP_WATCH_DIR, filename)
        fname_lower = filename.lower()
        
        if fname_lower.startswith('bom') or fname_lower.startswith('arbre'):
            import_type = 'bom'
        elif fname_lower.startswith('material') or fname_lower.startswith('matiere'):
            import_type = 'materials'
        elif fname_lower.startswith('recette') or fname_lower.startswith('recipe'):
            import_type = 'recipes'
        else:
            import_type = 'materials'
        
        try:
            with open(filepath, 'rb') as f:
                content = f.read()
            
            upload = UploadFile(filename=filename, file=BytesIO(content))
            
            if import_type == 'bom':
                result = await import_bom_csv(upload)
            elif import_type == 'recipes':
                result = await import_recipes_csv(upload)
            else:
                result = await import_materials_csv(upload)
            
            dest = os.path.join(processed_dir, f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}_{filename}")
            shutil.move(filepath, dest)
            
            log_entry = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "filename": filename,
                "import_type": import_type,
                "source": "sftp",
                "result": result
            }
            _append_import_log(log_entry)
            
            results.append({"filename": filename, "type": import_type, "result": result})
        except Exception as e:
            results.append({"filename": filename, "type": import_type, "error": str(e)})
    
    return {
        "files_scanned": len(files_found),
        "results": results,
        "watch_directory": SFTP_WATCH_DIR,
        "processed_directory": processed_dir
    }

@api_router.get("/import/status")
async def import_status():
    """Get import configuration and recent import history."""
    pending_files = [f for f in os.listdir(SFTP_WATCH_DIR) if f.endswith('.csv')]
    processed_dir = os.path.join(SFTP_WATCH_DIR, "processed")
    processed_files = []
    if os.path.exists(processed_dir):
        processed_files = [f for f in os.listdir(processed_dir) if f.endswith('.csv')]
    
    recent_logs = _read_import_logs(limit=20)
    
    return {
        "watch_directory": SFTP_WATCH_DIR,
        "pending_files": pending_files,
        "processed_count": len(processed_files),
        "recent_imports": recent_logs,
        "api_endpoints": {
            "auto_import": "POST /api/import/auto?import_type=materials|recipes|bom",
            "sftp_scan": "POST /api/import/sftp-scan",
            "status": "GET /api/import/status"
        }
    }

def _append_import_log(entry):
    """Save import log to MongoDB"""
    import asyncio
    entry["id"] = str(uuid.uuid4())
    if "timestamp" not in entry:
        entry["timestamp"] = datetime.now(timezone.utc).isoformat()
    try:
        loop = asyncio.get_event_loop()
        loop.create_task(db.import_logs.insert_one(entry))
    except:
        pass

def _read_import_logs(limit=20):
    """Read import logs - for sync context, return empty (use API endpoint)"""
    return []

@api_router.get("/import/logs")
async def get_import_logs(limit: int = 50):
    """Get import logs from database"""
    logs = await db.import_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs

# ================= PDF EXPORT =================

@api_router.get("/recipes/{recipe_id}/pdf")
async def export_recipe_pdf(recipe_id: str):
    cost = await calculate_cost(recipe_id)
    recipe = await db.recipes.find_one({"id": recipe_id}, {"_id": 0})

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#002FA7'), alignment=TA_CENTER, spaceAfter=10)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#09090B'), spaceBefore=16, spaceAfter=8)
    small_style = ParagraphStyle('SmallText', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#71717A'))

    table_base_style = [
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    elements = []
    elements.append(Paragraph("FICHE DE PRIX DE REVIENT", title_style))

    # Info block
    client_name = recipe.get('supplier_name', '-') or '-'
    product_type = recipe.get('product_type', '-') or '-'
    info_data = [
        ["Produit", cost.recipe_name, "Client", client_name],
        ["Production", f"{cost.output_quantity} {cost.output_unit}", "Type", product_type],
        ["Version", f"v{recipe.get('version', 1)}", "Date", datetime.now().strftime('%d/%m/%Y %H:%M')],
    ]
    if recipe.get('description'):
        info_data.append(["Description", recipe['description'], "", ""])
    info_table = Table(info_data, colWidths=[3*cm, 5.5*cm, 3*cm, 5.5*cm])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#71717A')),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#71717A')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E4E4E7')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))

    # Materials Table
    if cost.material_details:
        elements.append(Paragraph("1. MATIERES PREMIERES", heading_style))
        mat_data = [["Code", "Matiere", "Qte", "Unite", "Prix/u", "Freinte", "Total"]]
        for mat in cost.material_details:
            mat_data.append([mat.get('code_article', ''), mat['name'], f"{mat['quantity']}", mat['unit'], f"{mat['unit_price']:.2f} EUR",
                           f"{mat.get('freinte', 0)}%", f"{mat['total_cost']:.2f} EUR"])
        mat_data.append(["", "", "", "", "", "Sous-total:", f"{cost.total_material_cost:.2f} EUR"])
        mat_table = Table(mat_data, colWidths=[2*cm, 3.5*cm, 1.5*cm, 1.2*cm, 2.2*cm, 1.8*cm, 2.5*cm])
        mat_table.setStyle(TableStyle(table_base_style + [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#002FA7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F4F4F5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(mat_table)

    # Sub-recipes
    if cost.sub_recipe_details:
        elements.append(Paragraph("2. ARTICLES SEMI-FINIS", heading_style))
        for sub in cost.sub_recipe_details:
            sub_header = [["Semi-fini", "Qte", "Unite", "Cout/u", "Total"]]
            sub_header.append([sub['name'], f"{sub['quantity']}", sub['unit'],
                             f"{sub['unit_cost']:.2f} EUR", f"{sub['total_cost']:.2f} EUR"])
            sub_table = Table(sub_header, colWidths=[5*cm, 2*cm, 2*cm, 3*cm, 3*cm])
            sub_table.setStyle(TableStyle(table_base_style + [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            elements.append(sub_table)
            # Detail of sub-recipe ingredients
            if sub.get('ingredients'):
                detail_data = [["  Composition", "Qte", "Unite", "Prix/u", "Total"]]
                for si in sub['ingredients']:
                    detail_data.append([
                        f"  {si['name']}", f"{si['quantity']}", si.get('unit', ''),
                        f"{si.get('unit_price', 0):.2f} EUR" if not si.get('is_sub_recipe') else "-",
                        f"{si['total_cost']:.2f} EUR"
                    ])
                detail_table = Table(detail_data, colWidths=[5*cm, 2*cm, 2*cm, 3*cm, 3*cm])
                detail_table.setStyle(TableStyle([
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#71717A')),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E4E4E7')),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFBEB')),
                ]))
                elements.append(detail_table)
            elements.append(Spacer(1, 4))

    # Labor Table
    section_num = 3 if cost.sub_recipe_details else 2
    if cost.labor_details:
        elements.append(Paragraph(f"{section_num}. MAIN D'OEUVRE", heading_style))
        labor_data = [["Description", "Heures", "Taux horaire", "Total"]]
        for labor in cost.labor_details:
            labor_data.append([labor['description'], f"{labor['hours']} h", f"{labor['hourly_rate']:.2f} EUR/h", f"{labor['total_cost']:.2f} EUR"])
        labor_data.append(["", "", "Sous-total:", f"{cost.total_labor_cost:.2f} EUR"])
        labor_table = Table(labor_data, colWidths=[6*cm, 3*cm, 3.5*cm, 3*cm])
        labor_table.setStyle(TableStyle(table_base_style + [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F4F4F5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(labor_table)
        section_num += 1

    # Overheads
    if cost.overhead_details:
        elements.append(Paragraph(f"{section_num}. FRAIS GENERAUX", heading_style))
        oh_data = [["Frais", "Categorie", "Methode", "Total"]]
        for oh in cost.overhead_details:
            oh_data.append([oh['name'], oh['category'], oh['allocation_method'], f"{oh['total_cost']:.2f} EUR"])
        oh_data.append(["", "", "Sous-total:", f"{cost.total_overhead_cost:.2f} EUR"])
        oh_table = Table(oh_data, colWidths=[5*cm, 3.5*cm, 4*cm, 3*cm])
        oh_table.setStyle(TableStyle(table_base_style + [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F4F4F5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(oh_table)

    # Summary with margin
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("RECAPITULATIF ET MARGE", heading_style))
    summary_data = [
        ["Cout matieres (dont freinte)", f"{cost.total_material_cost:.2f} EUR"],
        ["Cout de main d'oeuvre", f"{cost.total_labor_cost:.2f} EUR"],
        ["Frais generaux", f"{cost.total_overhead_cost:.2f} EUR"],
        ["COUT TOTAL", f"{cost.total_cost:.2f} EUR"],
        [f"PRIX DE REVIENT / {cost.output_unit}", f"{cost.cost_per_unit:.2f} EUR"],
        [f"Marge cible ({cost.target_margin}%)", ""],
        ["PRIX DE VENTE CONSEILLE", f"{cost.suggested_price:.2f} EUR"],
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#002FA7')),
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.white),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#09090B')),
        ('TEXTCOLOR', (0, 4), (-1, 4), colors.white),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 6), (-1, 6), colors.white),
        ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 6), (-1, 6), 13),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Document genere le {datetime.now().strftime('%d/%m/%Y a %H:%M')} - PrixRevient", small_style))

    doc.build(elements)
    buffer.seek(0)

    filename = f"prix_revient_{cost.recipe_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

# ================= EXCEL EXPORT (XLSX) =================

@api_router.get("/reports/export-excel")
async def export_all_costs_xlsx():
    """Export all costs as real XLSX file"""
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau des couts"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["Recette", "Version", "Client", "Type", "Qte produite", "Unite", "Cout Matieres", "Cout Main oeuvre",
               "Cout Frais Gen.", "Cout Freinte", "Cout Total", "Prix/Unite", "Marge %", "Prix Conseille"]
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    
    row_num = 2
    money_fmt = '#,##0.00 "EUR"'
    
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe["id"])
            rtype = recipe.get("product_type", "") or ("SF" if recipe.get("is_intermediate") else "")
            values = [
                recipe["name"], f"v{recipe.get('version', 1)}", recipe.get("supplier_name", ""),
                rtype, recipe.get("output_quantity", 1), recipe.get("output_unit", "piece"),
                cost.total_material_cost, cost.total_labor_cost, cost.total_overhead_cost,
                cost.total_freinte_cost, cost.total_cost, cost.cost_per_unit,
                cost.target_margin, cost.suggested_price
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = border
                if col >= 7:
                    cell.number_format = money_fmt if col != 13 else '0.0"%"'
                    cell.alignment = Alignment(horizontal="right")
            
            if recipe.get("is_intermediate"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
            row_num += 1
        except:
            continue
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, row_num):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = min(max_len + 4, 30)
    
    # Total row
    if row_num > 2:
        total_row = row_num
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
        for col in [5, 6, 7, 8, 9]:
            cell = ws.cell(row=total_row, column=col)
            cell.value = f"=SUM({chr(64+col)}2:{chr(64+col)}{row_num-1})"
            cell.font = Font(bold=True)
            cell.number_format = money_fmt
            cell.border = border
            cell.fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapport_couts_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ================= UNITS MANAGEMENT =================

DEFAULT_UNITS = [
    {"id": "kg", "name": "Kilogramme", "abbreviation": "kg", "type": "poids"},
    {"id": "g", "name": "Gramme", "abbreviation": "g", "type": "poids"},
    {"id": "L", "name": "Litre", "abbreviation": "L", "type": "volume"},
    {"id": "mL", "name": "Millilitre", "abbreviation": "mL", "type": "volume"},
    {"id": "piece", "name": "Piece", "abbreviation": "pc", "type": "quantite"},
    {"id": "unite", "name": "Unite", "abbreviation": "u", "type": "quantite"},
    {"id": "lot", "name": "Lot", "abbreviation": "lot", "type": "quantite"},
    {"id": "boite", "name": "Boite", "abbreviation": "bte", "type": "quantite"},
    {"id": "m", "name": "Metre", "abbreviation": "m", "type": "longueur"},
    {"id": "cm", "name": "Centimetre", "abbreviation": "cm", "type": "longueur"},
]

@api_router.get("/units")
async def get_units():
    units = await db.units.find({}, {"_id": 0}).to_list(1000)
    if not units:
        # Seed default units
        for u in DEFAULT_UNITS:
            await db.units.insert_one({**u})
        units = await db.units.find({}, {"_id": 0}).to_list(1000)
    return units

@api_router.post("/units")
async def create_unit(request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    unit = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", ""),
        "abbreviation": data.get("abbreviation", ""),
        "type": data.get("type", "quantite"),
    }
    await db.units.insert_one(unit)
    return await db.units.find_one({"id": unit["id"]}, {"_id": 0})

@api_router.put("/units/{unit_id}")
async def update_unit(unit_id: str, request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    data.pop("id", None)
    data.pop("_id", None)
    await db.units.update_one({"id": unit_id}, {"$set": data})
    return await db.units.find_one({"id": unit_id}, {"_id": 0})

@api_router.delete("/units/{unit_id}")
async def delete_unit(unit_id: str, admin: dict = Depends(require_admin)):
    await db.units.delete_one({"id": unit_id})
    return {"message": "Unite supprimee"}

# ================= APP SETTINGS =================

DEFAULT_SETTINGS = {
    "primary_color": "#002FA7",
    "secondary_color": "#10B981",
    "accent_color": "#F59E0B",
    "danger_color": "#EF4444",
    "sidebar_bg": "#F4F4F5",
    "sidebar_text": "#71717A",
    "sidebar_active_bg": "#002FA7",
    "sidebar_active_text": "#FFFFFF",
    "company_name": "PrixRevient",
    "logo_data": "",
    "sso_enabled": False,
    "sso_google_enabled": False,
    "sso_google_client_id": "",
    "sso_google_client_secret": "",
    "sso_microsoft_enabled": False,
    "sso_microsoft_client_id": "",
    "sso_microsoft_client_secret": "",
    "sso_microsoft_tenant_id": "",
    "alert_threshold": 10,
}

@api_router.get("/settings")
async def get_settings():
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings:
        return DEFAULT_SETTINGS
    settings.pop("key", None)
    merged = {**DEFAULT_SETTINGS, **settings}
    return merged

@api_router.put("/settings")
async def update_settings(request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    data.pop("key", None)
    data.pop("_id", None)
    data["key"] = "app_settings"
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data["updated_by"] = admin["_id"]
    await db.settings.update_one({"key": "app_settings"}, {"$set": data}, upsert=True)
    result = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    result.pop("key", None)
    return {**DEFAULT_SETTINGS, **result}

@api_router.post("/settings/logo")
async def upload_logo(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Le fichier doit etre une image")
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image trop volumineuse (max 2 Mo)")
    b64 = base64.b64encode(content).decode("utf-8")
    logo_data = f"data:{file.content_type};base64,{b64}"
    await db.settings.update_one(
        {"key": "app_settings"},
        {"$set": {"logo_data": logo_data, "key": "app_settings"}},
        upsert=True
    )
    return {"logo_data": logo_data}

@api_router.delete("/settings/logo")
async def delete_logo(admin: dict = Depends(require_admin)):
    await db.settings.update_one({"key": "app_settings"}, {"$set": {"logo_data": ""}})
    return {"message": "Logo supprime"}

# ================= CRONTAB MANAGEMENT =================

@api_router.get("/crontabs")
async def get_crontabs(admin: dict = Depends(require_admin)):
    crontabs = await db.crontabs.find({}, {"_id": 0}).to_list(100)
    return crontabs

@api_router.post("/crontabs")
async def create_crontab(request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    crontab = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", "Import automatique"),
        "type": data.get("type", "sftp_scan"),
        "schedule": data.get("schedule", "*/30 * * * *"),
        "enabled": data.get("enabled", True),
        "last_run": None,
        "last_result": None,
        "last_status": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.crontabs.insert_one(crontab)
    crontab.pop("_id", None)
    await sync_scheduler()
    return crontab

@api_router.put("/crontabs/{crontab_id}")
async def update_crontab(crontab_id: str, request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    data.pop("_id", None)
    data.pop("id", None)
    await db.crontabs.update_one({"id": crontab_id}, {"$set": data})
    result = await db.crontabs.find_one({"id": crontab_id}, {"_id": 0})
    await sync_scheduler()
    return result

@api_router.delete("/crontabs/{crontab_id}")
async def delete_crontab(crontab_id: str, admin: dict = Depends(require_admin)):
    await db.crontabs.delete_one({"id": crontab_id})
    await sync_scheduler()
    return {"message": "Crontab supprime"}

@api_router.post("/crontabs/{crontab_id}/run")
async def run_crontab_now(crontab_id: str, admin: dict = Depends(require_admin)):
    crontab = await db.crontabs.find_one({"id": crontab_id}, {"_id": 0})
    if not crontab:
        raise HTTPException(status_code=404, detail="Crontab non trouve")
    
    result = await _execute_cron_task(crontab)
    return {"message": "Execute", "result": result}


async def _execute_cron_task(crontab: dict):
    """Execute a crontab task and update its status in DB."""
    task_type = crontab.get("type", "sftp_scan")
    crontab_id = crontab["id"]
    result_str = "OK"
    status = "success"
    try:
        if task_type == "sftp_scan":
            result = await sftp_scan()
            result_str = str(result) if result else "OK"
        elif task_type == "price_history":
            recipes = await db.recipes.find({}, {"_id": 0}).to_list(5000)
            recorded = 0
            for recipe in recipes:
                try:
                    cost = await calculate_cost(recipe["id"])
                    entry = {
                        "id": str(uuid.uuid4()),
                        "recipe_id": recipe["id"],
                        "recipe_name": recipe["name"],
                        "supplier_name": recipe.get("supplier_name", ""),
                        "version": recipe.get("version", ""),
                        "cost_per_unit": cost["cost_per_unit"],
                        "total_cost": cost["total_cost"],
                        "recorded_at": datetime.now(timezone.utc).isoformat(),
                    }
                    await db.price_history.insert_one(entry)
                    recorded += 1
                except Exception:
                    pass
            result_str = f"{recorded} recettes enregistrees"
        else:
            result_str = f"Type inconnu: {task_type}"
            status = "error"
    except Exception as e:
        result_str = str(e)
        status = "error"
    
    await db.crontabs.update_one({"id": crontab_id}, {"$set": {
        "last_run": datetime.now(timezone.utc).isoformat(),
        "last_result": result_str,
        "last_status": status,
    }})
    return result_str


# ================= BACKGROUND SCHEDULER =================

scheduler = BackgroundScheduler()
_scheduler_started = False


def _run_cron_in_loop(crontab_id: str):
    """Bridge from sync scheduler to async task execution."""
    import asyncio
    loop = asyncio.new_event_loop()
    try:
        loop.run_until_complete(_run_cron_async(crontab_id))
    finally:
        loop.close()


async def _run_cron_async(crontab_id: str):
    """Async cron execution with its own MongoDB client."""
    from motor.motor_asyncio import AsyncIOMotorClient
    mongo_url = os.environ.get("MONGO_URL", "")
    db_name = os.environ.get("DB_NAME", "cost_calculator")
    temp_client = AsyncIOMotorClient(mongo_url)
    temp_db = temp_client[db_name]
    try:
        crontab = await temp_db.crontabs.find_one({"id": crontab_id}, {"_id": 0})
        if not crontab or not crontab.get("enabled", False):
            return
        task_type = crontab.get("type", "sftp_scan")
        result_str = "OK"
        status = "success"
        try:
            if task_type == "sftp_scan":
                result_str = "SFTP scan execute"
            elif task_type == "price_history":
                recipes = await temp_db.recipes.find({}, {"_id": 0}).to_list(5000)
                recorded = 0
                for recipe in recipes:
                    try:
                        entry = {
                            "id": str(uuid.uuid4()),
                            "recipe_id": recipe["id"],
                            "recipe_name": recipe["name"],
                            "supplier_name": recipe.get("supplier_name", ""),
                            "version": recipe.get("version", ""),
                            "cost_per_unit": 0,
                            "total_cost": 0,
                            "recorded_at": datetime.now(timezone.utc).isoformat(),
                        }
                        await temp_db.price_history.insert_one(entry)
                        recorded += 1
                    except Exception:
                        pass
                result_str = f"{recorded} recettes enregistrees"
            else:
                result_str = f"Type inconnu: {task_type}"
                status = "error"
        except Exception as e:
            result_str = str(e)
            status = "error"
        await temp_db.crontabs.update_one({"id": crontab_id}, {"$set": {
            "last_run": datetime.now(timezone.utc).isoformat(),
            "last_result": result_str,
            "last_status": status,
        }})
    finally:
        temp_client.close()


def _parse_cron_schedule(schedule_str: str):
    """Parse cron expression to APScheduler CronTrigger."""
    parts = schedule_str.strip().split()
    if len(parts) != 5:
        return None
    return CronTrigger(
        minute=parts[0],
        hour=parts[1],
        day=parts[2],
        month=parts[3],
        day_of_week=parts[4],
    )


async def sync_scheduler():
    """Load all enabled crontabs from DB and schedule them."""
    global _scheduler_started
    
    # Remove all existing cron jobs
    for job in scheduler.get_jobs():
        if job.id.startswith("cron_"):
            scheduler.remove_job(job.id)
    
    crontabs = await db.crontabs.find({"enabled": True}, {"_id": 0}).to_list(100)
    for cron in crontabs:
        trigger = _parse_cron_schedule(cron.get("schedule", "*/30 * * * *"))
        if trigger:
            job_id = f"cron_{cron['id']}"
            scheduler.add_job(
                _run_cron_in_loop,
                trigger=trigger,
                id=job_id,
                args=[cron["id"]],
                replace_existing=True,
                misfire_grace_time=60,
            )
    
    if not _scheduler_started:
        scheduler.start()
        _scheduler_started = True
    
    logger.info(f"Scheduler synced: {len(crontabs)} tache(s) active(s)")


@api_router.get("/scheduler/status")
async def get_scheduler_status(admin: dict = Depends(require_admin)):
    """Get the scheduler status and next run times."""
    jobs = []
    for job in scheduler.get_jobs():
        if job.id.startswith("cron_"):
            cron_id = job.id.replace("cron_", "")
            jobs.append({
                "cron_id": cron_id,
                "next_run": job.next_run_time.isoformat() if job.next_run_time else None,
            })
    return {"running": _scheduler_started, "jobs": jobs}

# ================= DASHBOARD STATS =================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    total_materials = await db.raw_materials.count_documents({})
    total_recipes = await db.recipes.count_documents({})
    total_overheads = await db.overheads.count_documents({})
    total_suppliers = await db.suppliers.count_documents({})
    
    recipes = await db.recipes.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_recipes = []
    total_cost_sum = 0.0
    count = 0
    
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe['id'])
            recent_recipes.append({
                "id": recipe['id'],
                "name": recipe['name'],
                "cost_per_unit": cost.cost_per_unit,
                "suggested_price": cost.suggested_price,
                "output_unit": recipe.get('output_unit', 'pièce')
            })
            total_cost_sum += cost.cost_per_unit
            count += 1
        except:
            continue
    
    avg_cost = total_cost_sum / count if count > 0 else 0.0
    
    # Stock alerts count
    stock_alerts = await db.raw_materials.count_documents({
        "$expr": {"$lte": ["$stock_quantity", "$stock_alert_threshold"]},
        "stock_alert_threshold": {"$gt": 0}
    })
    
    return {
        "total_materials": total_materials,
        "total_recipes": total_recipes,
        "total_overheads": total_overheads,
        "total_suppliers": total_suppliers,
        "avg_cost_per_unit": round(avg_cost, 2),
        "recent_recipes": recent_recipes,
        "stock_alerts": stock_alerts
    }

@api_router.get("/dashboard/admin-stats")
async def get_admin_stats(admin: dict = Depends(require_admin)):
    """Admin-only dashboard statistics"""
    # Users by role
    all_users = await db.users.find({}, {"_id": 0, "role": 1, "is_active": 1, "email": 1, "name": 1, "created_at": 1}).to_list(1000)
    users_by_role = {}
    active_count = 0
    for u in all_users:
        r = u.get("role", "operator")
        users_by_role[r] = users_by_role.get(r, 0) + 1
        if u.get("is_active", True):
            active_count += 1

    # Recent imports (last 10)
    recent_imports = await db.import_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(10)

    # Import stats
    total_imports = await db.import_logs.count_documents({})
    success_imports = await db.import_logs.count_documents({"status": "success"})
    error_imports = await db.import_logs.count_documents({"status": "error"})

    # Recipes by category
    recipes_pipeline = [
        {"$group": {"_id": "$category_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    recipes_by_cat_raw = await db.recipes.aggregate(recipes_pipeline).to_list(100)
    cat_ids = [r["_id"] for r in recipes_by_cat_raw if r["_id"]]
    categories = {}
    if cat_ids:
        cats = await db.categories.find({"id": {"$in": cat_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        categories = {c["id"]: c["name"] for c in cats}
    recipes_by_category = [
        {"category": categories.get(r["_id"], "Non classees"), "count": r["count"]}
        for r in recipes_by_cat_raw
    ]

    # Materials by category
    materials_pipeline = [
        {"$group": {"_id": "$category_id", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    materials_by_cat_raw = await db.raw_materials.aggregate(materials_pipeline).to_list(100)
    materials_by_category = [
        {"category": categories.get(r["_id"], "Non classees"), "count": r["count"]}
        for r in materials_by_cat_raw
    ]

    # Crontab statuses
    crontabs = await db.crontabs.find({}, {"_id": 0}).to_list(50)
    crontab_summary = [
        {"name": c.get("name", ""), "type": c.get("type", ""), "enabled": c.get("enabled", False),
         "last_status": c.get("last_status", ""), "last_run": c.get("last_run", "")}
        for c in crontabs
    ]

    # Sites count
    total_sites = await db.sites.count_documents({})

    # Stock alerts details (top 5)
    low_stock = await db.raw_materials.find(
        {"$expr": {"$lte": ["$stock_quantity", "$stock_alert_threshold"]}, "stock_alert_threshold": {"$gt": 0}},
        {"_id": 0, "name": 1, "stock_quantity": 1, "stock_alert_threshold": 1, "unit": 1}
    ).limit(5).to_list(5)

    return {
        "total_users": len(all_users),
        "active_users": active_count,
        "users_by_role": users_by_role,
        "total_imports": total_imports,
        "success_imports": success_imports,
        "error_imports": error_imports,
        "recent_imports": recent_imports,
        "recipes_by_category": recipes_by_category,
        "materials_by_category": materials_by_category,
        "crontab_summary": crontab_summary,
        "total_sites": total_sites,
        "low_stock_items": low_stock,
    }



# ================= DATA MANAGEMENT (SEED / RESET / QUERY) =================

SEED_CATEGORIES = [
    {"name": "Farines & cereales", "description": "Farine, semoule, amidon, flocons"},
    {"name": "Produits laitiers", "description": "Lait, creme, beurre, fromage"},
    {"name": "Fruits & legumes", "description": "Fruits frais, legumes, purees"},
    {"name": "Sucres & edulcorants", "description": "Sucre, miel, sirop, glucose"},
    {"name": "Matieres grasses", "description": "Huiles, margarines, saindoux"},
    {"name": "Epices & aromes", "description": "Vanille, cannelle, extraits, colorants"},
    {"name": "Oeufs & ovoproduits", "description": "Oeufs entiers, blancs, jaunes"},
    {"name": "Chocolat & cacao", "description": "Chocolat noir, lait, blanc, cacao"},
]

SEED_SUPPLIERS = [
    {"name": "Carrefour", "code": "CLI-001", "contact": "Jean Dupont", "email": "achat@carrefour.fr", "phone": "0321456789", "address": "12 rue du Commerce, 59000 Lille"},
    {"name": "Leclerc", "code": "CLI-002", "contact": "Marie Martin", "email": "pro@leclerc.fr", "phone": "0322334455", "address": "45 avenue Leclerc, 62000 Arras"},
    {"name": "Auchan", "code": "CLI-003", "contact": "Pierre Leroy", "email": "commandes@auchan.fr", "phone": "0145678901", "address": "8 centre commercial, 94150 Rungis"},
    {"name": "Intermarche", "code": "CLI-004", "contact": "Sophie Moreau", "email": "ventes@intermarche.fr", "phone": "0380123456", "address": "Route de Dijon, 21000 Dijon"},
    {"name": "Lidl", "code": "CLI-005", "contact": "Laurent Blanc", "email": "pro@lidl.fr", "phone": "0478901234", "address": "22 rue du Discount, 69000 Lyon"},
    {"name": "Metro", "code": "CLI-006", "contact": "Fatima Amrani", "email": "info@metro.fr", "phone": "0491234567", "address": "15 quai des Pros, 13002 Marseille"},
]

SEED_UNITS = [
    {"name": "Kilogramme", "abbreviation": "kg", "type": "poids"},
    {"name": "Gramme", "abbreviation": "g", "type": "poids"},
    {"name": "Litre", "abbreviation": "L", "type": "volume"},
    {"name": "Millilitre", "abbreviation": "mL", "type": "volume"},
    {"name": "Piece", "abbreviation": "pce", "type": "quantite"},
    {"name": "Metre", "abbreviation": "m", "type": "longueur"},
]

SEED_OVERHEADS = [
    {"name": "Electricite atelier", "category": "Energie", "monthly_amount": 850.0, "allocation_method": "per_unit", "allocation_value": 500},
    {"name": "Eau", "category": "Energie", "monthly_amount": 180.0, "allocation_method": "per_unit", "allocation_value": 500},
    {"name": "Location atelier", "category": "Infrastructure", "monthly_amount": 2200.0, "allocation_method": "per_hour", "allocation_value": 160},
    {"name": "Maintenance equipement", "category": "Equipement", "monthly_amount": 350.0, "allocation_method": "per_unit", "allocation_value": 500},
    {"name": "Assurance professionnelle", "category": "Infrastructure", "monthly_amount": 420.0, "allocation_method": "per_unit", "allocation_value": 500},
]

SEED_MATERIALS_SPEC = [
    {"name": "Farine de ble T55", "code_article": "MAT-001", "unit": "kg", "unit_price": 1.20, "category": "Farines & cereales", "supplier": "Carrefour", "freinte": 2.0, "stock_quantity": 150, "stock_alert_threshold": 20},
    {"name": "Farine de ble T65", "code_article": "MAT-002", "unit": "kg", "unit_price": 1.35, "category": "Farines & cereales", "supplier": "Carrefour", "freinte": 2.0, "stock_quantity": 80, "stock_alert_threshold": 15},
    {"name": "Beurre doux 82%", "code_article": "MAT-003", "unit": "kg", "unit_price": 8.50, "category": "Produits laitiers", "supplier": "Leclerc", "freinte": 1.0, "stock_quantity": 25, "stock_alert_threshold": 5},
    {"name": "Sucre semoule", "code_article": "MAT-004", "unit": "kg", "unit_price": 1.10, "category": "Sucres & edulcorants", "supplier": "Intermarche", "freinte": 0.5, "stock_quantity": 100, "stock_alert_threshold": 15},
    {"name": "Oeufs frais (calibre M)", "code_article": "MAT-005", "unit": "pce", "unit_price": 0.28, "category": "Oeufs & ovoproduits", "supplier": "Leclerc", "freinte": 3.0, "stock_quantity": 360, "stock_alert_threshold": 60},
    {"name": "Lait entier", "code_article": "MAT-006", "unit": "L", "unit_price": 1.05, "category": "Produits laitiers", "supplier": "Leclerc", "freinte": 1.0, "stock_quantity": 40, "stock_alert_threshold": 10},
    {"name": "Creme fraiche 35%", "code_article": "MAT-007", "unit": "L", "unit_price": 4.80, "category": "Produits laitiers", "supplier": "Leclerc", "freinte": 2.0, "stock_quantity": 15, "stock_alert_threshold": 5},
    {"name": "Pommes Golden", "code_article": "MAT-008", "unit": "kg", "unit_price": 2.90, "category": "Fruits & legumes", "supplier": "Auchan", "freinte": 8.0, "stock_quantity": 30, "stock_alert_threshold": 10},
    {"name": "Chocolat noir 70%", "code_article": "MAT-009", "unit": "kg", "unit_price": 12.50, "category": "Chocolat & cacao", "supplier": "Lidl", "freinte": 1.5, "stock_quantity": 10, "stock_alert_threshold": 3},
    {"name": "Levure boulangere", "code_article": "MAT-010", "unit": "kg", "unit_price": 5.20, "category": "Farines & cereales", "supplier": "Carrefour", "freinte": 0.0, "stock_quantity": 5, "stock_alert_threshold": 1},
    {"name": "Extrait de vanille", "code_article": "MAT-011", "unit": "L", "unit_price": 85.00, "category": "Epices & aromes", "supplier": "Metro", "freinte": 0.0, "stock_quantity": 2, "stock_alert_threshold": 0.5},
    {"name": "Sel fin", "code_article": "MAT-012", "unit": "kg", "unit_price": 0.65, "category": "Epices & aromes", "supplier": "Carrefour", "freinte": 0.0, "stock_quantity": 25, "stock_alert_threshold": 5},
    {"name": "Poudre d'amandes", "code_article": "MAT-013", "unit": "kg", "unit_price": 14.80, "category": "Fruits & legumes", "supplier": "Auchan", "freinte": 1.0, "stock_quantity": 8, "stock_alert_threshold": 2},
    {"name": "Sucre glace", "code_article": "MAT-014", "unit": "kg", "unit_price": 2.10, "category": "Sucres & edulcorants", "supplier": "Intermarche", "freinte": 1.0, "stock_quantity": 20, "stock_alert_threshold": 5},
    {"name": "Cacao en poudre", "code_article": "MAT-015", "unit": "kg", "unit_price": 9.80, "category": "Chocolat & cacao", "supplier": "Lidl", "freinte": 0.5, "stock_quantity": 6, "stock_alert_threshold": 2},
    {"name": "Huile de tournesol", "code_article": "MAT-016", "unit": "L", "unit_price": 2.40, "category": "Matieres grasses", "supplier": "Intermarche", "freinte": 0.5, "stock_quantity": 20, "stock_alert_threshold": 5},
    {"name": "Miel toutes fleurs", "code_article": "MAT-017", "unit": "kg", "unit_price": 11.50, "category": "Sucres & edulcorants", "supplier": "Metro", "freinte": 0.0, "stock_quantity": 5, "stock_alert_threshold": 1},
    {"name": "Poires Williams", "code_article": "MAT-018", "unit": "kg", "unit_price": 3.50, "category": "Fruits & legumes", "supplier": "Auchan", "freinte": 10.0, "stock_quantity": 15, "stock_alert_threshold": 5},
]

@api_router.post("/data/seed")
async def seed_data(admin: dict = Depends(require_admin)):
    """Seed the database with sample data"""
    now = datetime.now(timezone.utc)
    counts = {}

    # Categories
    cat_map = {}
    inserted_cats = 0
    for c in SEED_CATEGORIES:
        exists = await db.categories.find_one({"name": c["name"]})
        if not exists:
            cid = str(uuid.uuid4())
            await db.categories.insert_one({"id": cid, "name": c["name"], "description": c["description"], "created_at": now})
            cat_map[c["name"]] = cid
            inserted_cats += 1
        else:
            cat_map[c["name"]] = exists["id"]
    counts["categories"] = inserted_cats

    # Suppliers
    sup_map = {}
    inserted_sups = 0
    for s in SEED_SUPPLIERS:
        exists = await db.suppliers.find_one({"name": s["name"]})
        if not exists:
            sid = str(uuid.uuid4())
            await db.suppliers.insert_one({"id": sid, **s, "created_at": now})
            sup_map[s["name"]] = sid
            inserted_sups += 1
        else:
            sup_map[s["name"]] = exists["id"]
    counts["suppliers"] = inserted_sups

    # Units
    inserted_units = 0
    for u in SEED_UNITS:
        exists = await db.units.find_one({"abbreviation": u["abbreviation"]})
        if not exists:
            await db.units.insert_one({"id": str(uuid.uuid4()), **u})
            inserted_units += 1
    counts["units"] = inserted_units

    # Overheads
    oh_ids = []
    inserted_oh = 0
    for o in SEED_OVERHEADS:
        exists = await db.overheads.find_one({"name": o["name"]})
        if not exists:
            oid = str(uuid.uuid4())
            await db.overheads.insert_one({"id": oid, **o, "created_at": now})
            oh_ids.append(oid)
            inserted_oh += 1
        else:
            oh_ids.append(exists["id"])
    counts["overheads"] = inserted_oh

    # Raw Materials
    mat_map = {}
    inserted_mats = 0
    for m in SEED_MATERIALS_SPEC:
        exists = await db.raw_materials.find_one({"name": m["name"]})
        if not exists:
            mid = str(uuid.uuid4())
            cat_id = cat_map.get(m["category"], "")
            sup_id = sup_map.get(m["supplier"], "")
            await db.raw_materials.insert_one({
                "id": mid, "name": m["name"], "code_article": m["code_article"],
                "unit": m["unit"], "unit_price": m["unit_price"],
                "supplier_id": sup_id, "supplier_name": m["supplier"],
                "category_id": cat_id, "description": "",
                "freinte": m["freinte"], "stock_quantity": m["stock_quantity"],
                "stock_alert_threshold": m["stock_alert_threshold"], "created_at": now
            })
            mat_map[m["name"]] = {"id": mid, "unit": m["unit"], "unit_price": m["unit_price"], "freinte": m["freinte"]}
            inserted_mats += 1
        else:
            mat_map[m["name"]] = {"id": exists["id"], "unit": exists.get("unit", "kg"), "unit_price": exists.get("unit_price", 0), "freinte": exists.get("freinte", 0)}
    counts["raw_materials"] = inserted_mats

    # Recipes
    inserted_recipes = 0

    # --- Pate brisee (intermediate) ---
    if not await db.recipes.find_one({"name": "Pate brisee"}):
        pb_id = str(uuid.uuid4())
        await db.recipes.insert_one({
            "id": pb_id, "name": "Pate brisee", "description": "Pate brisee classique pour tartes",
            "category_id": cat_map.get("Farines & cereales", ""), "supplier_id": "", "supplier_name": "",
            "version": 1, "output_quantity": 1.0, "output_unit": "kg", "target_margin": 30.0,
            "is_intermediate": True,
            "ingredients": [
                {"material_id": mat_map.get("Farine de ble T55", {}).get("id", ""), "material_name": "Farine de ble T55", "sub_recipe_id": None, "quantity": 0.5, "unit": "kg", "unit_price": 1.20, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Beurre doux 82%", {}).get("id", ""), "material_name": "Beurre doux 82%", "sub_recipe_id": None, "quantity": 0.25, "unit": "kg", "unit_price": 8.50, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Oeufs frais (calibre M)", {}).get("id", ""), "material_name": "Oeufs frais (calibre M)", "sub_recipe_id": None, "quantity": 1, "unit": "pce", "unit_price": 0.28, "freinte": 3.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sel fin", {}).get("id", ""), "material_name": "Sel fin", "sub_recipe_id": None, "quantity": 0.005, "unit": "kg", "unit_price": 0.65, "freinte": 0.0, "is_sub_recipe": False},
            ],
            "labor_costs": [{"description": "Petrissage et repos", "hours": 0.5, "hourly_rate": 15.0}],
            "overhead_ids": [oh_ids[0]] if oh_ids else [],
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1
    pb_doc = await db.recipes.find_one({"name": "Pate brisee"})
    pb_id = pb_doc["id"] if pb_doc else ""

    # --- Creme patissiere (intermediate) ---
    if not await db.recipes.find_one({"name": "Creme patissiere"}):
        cp_id = str(uuid.uuid4())
        await db.recipes.insert_one({
            "id": cp_id, "name": "Creme patissiere", "description": "Creme patissiere vanille onctueuse",
            "category_id": cat_map.get("Produits laitiers", ""), "supplier_id": "", "supplier_name": "",
            "version": 1, "output_quantity": 1.0, "output_unit": "kg", "target_margin": 30.0,
            "is_intermediate": True,
            "ingredients": [
                {"material_id": mat_map.get("Lait entier", {}).get("id", ""), "material_name": "Lait entier", "sub_recipe_id": None, "quantity": 0.5, "unit": "L", "unit_price": 1.05, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sucre semoule", {}).get("id", ""), "material_name": "Sucre semoule", "sub_recipe_id": None, "quantity": 0.1, "unit": "kg", "unit_price": 1.10, "freinte": 0.5, "is_sub_recipe": False},
                {"material_id": mat_map.get("Oeufs frais (calibre M)", {}).get("id", ""), "material_name": "Oeufs frais (calibre M)", "sub_recipe_id": None, "quantity": 3, "unit": "pce", "unit_price": 0.28, "freinte": 3.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Farine de ble T55", {}).get("id", ""), "material_name": "Farine de ble T55", "sub_recipe_id": None, "quantity": 0.04, "unit": "kg", "unit_price": 1.20, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Extrait de vanille", {}).get("id", ""), "material_name": "Extrait de vanille", "sub_recipe_id": None, "quantity": 0.005, "unit": "L", "unit_price": 85.0, "freinte": 0.0, "is_sub_recipe": False},
            ],
            "labor_costs": [{"description": "Cuisson et refroidissement", "hours": 0.75, "hourly_rate": 15.0}],
            "overhead_ids": [oh_ids[0]] if oh_ids else [],
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1
    cp_doc = await db.recipes.find_one({"name": "Creme patissiere"})
    cp_id = cp_doc["id"] if cp_doc else ""

    # --- Tarte aux pommes (final - uses Pate brisee) ---
    if not await db.recipes.find_one({"name": "Tarte aux pommes"}):
        await db.recipes.insert_one({
            "id": str(uuid.uuid4()), "name": "Tarte aux pommes", "description": "Tarte aux pommes tradition avec pate brisee maison",
            "category_id": cat_map.get("Fruits & legumes", ""), "supplier_id": sup_map.get("Auchan", ""), "supplier_name": "Auchan",
            "version": 1, "output_quantity": 8.0, "output_unit": "part", "target_margin": 35.0,
            "is_intermediate": False, "product_type": "MDD",
            "ingredients": [
                {"material_id": None, "material_name": "Pate brisee", "sub_recipe_id": pb_id, "quantity": 0.4, "unit": "kg", "unit_price": 0, "freinte": 0, "is_sub_recipe": True},
                {"material_id": mat_map.get("Pommes Golden", {}).get("id", ""), "material_name": "Pommes Golden", "sub_recipe_id": None, "quantity": 0.8, "unit": "kg", "unit_price": 2.90, "freinte": 8.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sucre semoule", {}).get("id", ""), "material_name": "Sucre semoule", "sub_recipe_id": None, "quantity": 0.08, "unit": "kg", "unit_price": 1.10, "freinte": 0.5, "is_sub_recipe": False},
                {"material_id": mat_map.get("Beurre doux 82%", {}).get("id", ""), "material_name": "Beurre doux 82%", "sub_recipe_id": None, "quantity": 0.03, "unit": "kg", "unit_price": 8.50, "freinte": 1.0, "is_sub_recipe": False},
            ],
            "labor_costs": [
                {"description": "Preparation et decoupe", "hours": 0.5, "hourly_rate": 15.0},
                {"description": "Cuisson et finition", "hours": 0.75, "hourly_rate": 15.0},
            ],
            "overhead_ids": oh_ids[:2] if len(oh_ids) >= 2 else oh_ids,
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1

    # --- Eclair au chocolat (uses Creme patissiere) ---
    if not await db.recipes.find_one({"name": "Eclair au chocolat"}):
        await db.recipes.insert_one({
            "id": str(uuid.uuid4()), "name": "Eclair au chocolat", "description": "Eclair garni de creme patissiere, glace chocolat",
            "category_id": cat_map.get("Chocolat & cacao", ""), "supplier_id": sup_map.get("Lidl", ""), "supplier_name": "Lidl",
            "version": 1, "output_quantity": 12.0, "output_unit": "piece", "target_margin": 40.0,
            "is_intermediate": False, "product_type": "MN",
            "ingredients": [
                {"material_id": mat_map.get("Farine de ble T55", {}).get("id", ""), "material_name": "Farine de ble T55", "sub_recipe_id": None, "quantity": 0.15, "unit": "kg", "unit_price": 1.20, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Beurre doux 82%", {}).get("id", ""), "material_name": "Beurre doux 82%", "sub_recipe_id": None, "quantity": 0.1, "unit": "kg", "unit_price": 8.50, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Oeufs frais (calibre M)", {}).get("id", ""), "material_name": "Oeufs frais (calibre M)", "sub_recipe_id": None, "quantity": 4, "unit": "pce", "unit_price": 0.28, "freinte": 3.0, "is_sub_recipe": False},
                {"material_id": None, "material_name": "Creme patissiere", "sub_recipe_id": cp_id, "quantity": 0.5, "unit": "kg", "unit_price": 0, "freinte": 0, "is_sub_recipe": True},
                {"material_id": mat_map.get("Chocolat noir 70%", {}).get("id", ""), "material_name": "Chocolat noir 70%", "sub_recipe_id": None, "quantity": 0.15, "unit": "kg", "unit_price": 12.50, "freinte": 1.5, "is_sub_recipe": False},
                {"material_id": mat_map.get("Creme fraiche 35%", {}).get("id", ""), "material_name": "Creme fraiche 35%", "sub_recipe_id": None, "quantity": 0.1, "unit": "L", "unit_price": 4.80, "freinte": 2.0, "is_sub_recipe": False},
            ],
            "labor_costs": [
                {"description": "Pate a choux + pochage", "hours": 1.0, "hourly_rate": 15.0},
                {"description": "Glacage et assemblage", "hours": 0.5, "hourly_rate": 15.0},
            ],
            "overhead_ids": oh_ids[:3] if len(oh_ids) >= 3 else oh_ids,
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1

    # --- Pain de campagne ---
    if not await db.recipes.find_one({"name": "Pain de campagne"}):
        await db.recipes.insert_one({
            "id": str(uuid.uuid4()), "name": "Pain de campagne", "description": "Pain rustique au levain avec melange de farines",
            "category_id": cat_map.get("Farines & cereales", ""), "supplier_id": sup_map.get("Carrefour", ""), "supplier_name": "Carrefour",
            "version": 1, "output_quantity": 4.0, "output_unit": "piece", "target_margin": 25.0,
            "is_intermediate": False, "product_type": "SM",
            "ingredients": [
                {"material_id": mat_map.get("Farine de ble T65", {}).get("id", ""), "material_name": "Farine de ble T65", "sub_recipe_id": None, "quantity": 1.0, "unit": "kg", "unit_price": 1.35, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Levure boulangere", {}).get("id", ""), "material_name": "Levure boulangere", "sub_recipe_id": None, "quantity": 0.02, "unit": "kg", "unit_price": 5.20, "freinte": 0.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sel fin", {}).get("id", ""), "material_name": "Sel fin", "sub_recipe_id": None, "quantity": 0.02, "unit": "kg", "unit_price": 0.65, "freinte": 0.0, "is_sub_recipe": False},
            ],
            "labor_costs": [
                {"description": "Petrissage et pousse", "hours": 0.5, "hourly_rate": 14.0},
                {"description": "Faconnage et cuisson", "hours": 1.0, "hourly_rate": 14.0},
            ],
            "overhead_ids": oh_ids[:2] if len(oh_ids) >= 2 else oh_ids,
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1

    # --- Croissant au beurre ---
    if not await db.recipes.find_one({"name": "Croissant au beurre"}):
        await db.recipes.insert_one({
            "id": str(uuid.uuid4()), "name": "Croissant au beurre", "description": "Croissant pur beurre feuillete",
            "category_id": cat_map.get("Produits laitiers", ""), "supplier_id": sup_map.get("Leclerc", ""), "supplier_name": "Leclerc",
            "version": 1, "output_quantity": 10.0, "output_unit": "piece", "target_margin": 45.0,
            "is_intermediate": False, "product_type": "MDD",
            "ingredients": [
                {"material_id": mat_map.get("Farine de ble T55", {}).get("id", ""), "material_name": "Farine de ble T55", "sub_recipe_id": None, "quantity": 0.5, "unit": "kg", "unit_price": 1.20, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Beurre doux 82%", {}).get("id", ""), "material_name": "Beurre doux 82%", "sub_recipe_id": None, "quantity": 0.25, "unit": "kg", "unit_price": 8.50, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sucre semoule", {}).get("id", ""), "material_name": "Sucre semoule", "sub_recipe_id": None, "quantity": 0.06, "unit": "kg", "unit_price": 1.10, "freinte": 0.5, "is_sub_recipe": False},
                {"material_id": mat_map.get("Lait entier", {}).get("id", ""), "material_name": "Lait entier", "sub_recipe_id": None, "quantity": 0.15, "unit": "L", "unit_price": 1.05, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Levure boulangere", {}).get("id", ""), "material_name": "Levure boulangere", "sub_recipe_id": None, "quantity": 0.015, "unit": "kg", "unit_price": 5.20, "freinte": 0.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sel fin", {}).get("id", ""), "material_name": "Sel fin", "sub_recipe_id": None, "quantity": 0.01, "unit": "kg", "unit_price": 0.65, "freinte": 0.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Oeufs frais (calibre M)", {}).get("id", ""), "material_name": "Oeufs frais (calibre M)", "sub_recipe_id": None, "quantity": 1, "unit": "pce", "unit_price": 0.28, "freinte": 3.0, "is_sub_recipe": False},
            ],
            "labor_costs": [
                {"description": "Petrissage et tourage", "hours": 1.5, "hourly_rate": 16.0},
                {"description": "Faconnage et cuisson", "hours": 0.75, "hourly_rate": 16.0},
            ],
            "overhead_ids": oh_ids[:3] if len(oh_ids) >= 3 else oh_ids,
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1

    # --- Financier aux amandes ---
    if not await db.recipes.find_one({"name": "Financier aux amandes"}):
        await db.recipes.insert_one({
            "id": str(uuid.uuid4()), "name": "Financier aux amandes", "description": "Petit gateau moelleux aux amandes et beurre noisette",
            "category_id": cat_map.get("Fruits & legumes", ""), "supplier_id": sup_map.get("Intermarche", ""), "supplier_name": "Intermarche",
            "version": 1, "output_quantity": 20.0, "output_unit": "piece", "target_margin": 50.0,
            "is_intermediate": False, "product_type": "MP",
            "ingredients": [
                {"material_id": mat_map.get("Poudre d'amandes", {}).get("id", ""), "material_name": "Poudre d'amandes", "sub_recipe_id": None, "quantity": 0.12, "unit": "kg", "unit_price": 14.80, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Sucre glace", {}).get("id", ""), "material_name": "Sucre glace", "sub_recipe_id": None, "quantity": 0.15, "unit": "kg", "unit_price": 2.10, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Beurre doux 82%", {}).get("id", ""), "material_name": "Beurre doux 82%", "sub_recipe_id": None, "quantity": 0.12, "unit": "kg", "unit_price": 8.50, "freinte": 1.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Oeufs frais (calibre M)", {}).get("id", ""), "material_name": "Oeufs frais (calibre M)", "sub_recipe_id": None, "quantity": 4, "unit": "pce", "unit_price": 0.28, "freinte": 3.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Farine de ble T55", {}).get("id", ""), "material_name": "Farine de ble T55", "sub_recipe_id": None, "quantity": 0.05, "unit": "kg", "unit_price": 1.20, "freinte": 2.0, "is_sub_recipe": False},
                {"material_id": mat_map.get("Miel toutes fleurs", {}).get("id", ""), "material_name": "Miel toutes fleurs", "sub_recipe_id": None, "quantity": 0.02, "unit": "kg", "unit_price": 11.50, "freinte": 0.0, "is_sub_recipe": False},
            ],
            "labor_costs": [{"description": "Preparation et cuisson", "hours": 0.75, "hourly_rate": 15.0}],
            "overhead_ids": oh_ids[:2] if len(oh_ids) >= 2 else oh_ids,
            "user_id": "admin@example.com", "created_at": now, "updated_at": now
        })
        inserted_recipes += 1

    counts["recipes"] = inserted_recipes

    return {"message": "Jeu de donnees charge avec succes", "inserted": counts}


@api_router.post("/data/reset")
async def reset_data(admin: dict = Depends(require_admin)):
    """Reset all data collections (keeps users and settings)"""
    collections_to_clear = [
        "raw_materials", "recipes", "categories", "suppliers",
        "overheads", "units", "sites", "crontabs",
        "import_logs", "price_history", "price_history_materials", "api_keys"
    ]
    deleted = {}
    for col_name in collections_to_clear:
        col = db[col_name]
        result = await col.delete_many({})
        deleted[col_name] = result.deleted_count

    # Re-create default site
    await db.sites.insert_one({
        "id": str(uuid.uuid4()), "name": "Site principal", "address": "",
        "is_default": True, "created_at": datetime.now(timezone.utc)
    })

    return {"message": "Donnees reinitialisees", "deleted": deleted}


@api_router.post("/data/query")
async def execute_query(request: Request, admin: dict = Depends(require_admin)):
    """Execute a MongoDB query (find/count/aggregate) - admin only"""
    import json as json_mod
    body = await request.json()
    collection_name = body.get("collection", "")
    operation = body.get("operation", "find")
    query_filter = body.get("filter", {})
    projection = body.get("projection", {})
    limit = min(body.get("limit", 50), 200)
    sort_field = body.get("sort", None)
    pipeline = body.get("pipeline", [])

    valid_collections = [
        "users", "raw_materials", "recipes", "categories", "suppliers",
        "overheads", "units", "sites", "crontabs", "import_logs",
        "price_history", "price_history_materials", "api_keys", "settings"
    ]
    if collection_name not in valid_collections:
        raise HTTPException(status_code=400, detail=f"Collection invalide. Collections: {', '.join(valid_collections)}")

    col = db[collection_name]

    # Always exclude _id and password_hash
    if "_id" not in projection:
        projection["_id"] = 0
    if collection_name == "users":
        projection["password_hash"] = 0

    try:
        if operation == "find":
            cursor = col.find(query_filter, projection)
            if sort_field:
                cursor = cursor.sort(sort_field, -1)
            results = await cursor.limit(limit).to_list(limit)
            return {"count": len(results), "results": results}

        elif operation == "count":
            count = await col.count_documents(query_filter)
            return {"count": count, "results": []}

        elif operation == "aggregate":
            results = await col.aggregate(pipeline).to_list(limit)
            for r in results:
                if "_id" in r and isinstance(r["_id"], ObjectId):
                    r["_id"] = str(r["_id"])
            return {"count": len(results), "results": results}

        elif operation == "distinct":
            field = body.get("field", "name")
            results = await col.distinct(field, query_filter)
            return {"count": len(results), "results": results}

        else:
            raise HTTPException(status_code=400, detail="Operation invalide. Utilisez: find, count, aggregate, distinct")

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur requete: {str(e)}")



# ================= ROOT =================

@api_router.get("/")
async def root():
    return {"message": "API Calculateur de Prix de Revient v2.0"}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ================= SSO OAUTH2 =================

@api_router.get("/auth/sso/google/url")
async def google_sso_url():
    """Get Google OAuth2 authorization URL"""
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings or not settings.get("sso_google_enabled") or not settings.get("sso_google_client_id"):
        raise HTTPException(status_code=400, detail="SSO Google non configure")
    client_id = settings["sso_google_client_id"]
    redirect_uri = os.environ.get("SSO_REDIRECT_URL", "") + "/api/auth/sso/google/callback"
    scope = "openid email profile"
    url = f"https://accounts.google.com/o/oauth2/v2/auth?client_id={client_id}&redirect_uri={redirect_uri}&response_type=code&scope={scope}&access_type=offline"
    return {"url": url}

@api_router.get("/auth/sso/google/callback")
async def google_sso_callback(code: str, response: Response):
    """Handle Google OAuth2 callback"""
    import httpx
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings or not settings.get("sso_google_enabled"):
        raise HTTPException(status_code=400, detail="SSO Google non active")
    
    client_id = settings["sso_google_client_id"]
    client_secret = settings.get("sso_google_client_secret", "")
    redirect_uri = os.environ.get("SSO_REDIRECT_URL", "") + "/api/auth/sso/google/callback"
    
    async with httpx.AsyncClient() as client:
        token_resp = await client.post("https://oauth2.googleapis.com/token", data={
            "code": code, "client_id": client_id, "client_secret": client_secret,
            "redirect_uri": redirect_uri, "grant_type": "authorization_code"
        })
        if token_resp.status_code != 200:
            raise HTTPException(status_code=400, detail="Erreur d'authentification Google")
        tokens = token_resp.json()
        
        userinfo_resp = await client.get("https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {tokens['access_token']}"})
        userinfo = userinfo_resp.json()
    
    email = userinfo.get("email", "")
    name = userinfo.get("name", email)
    
    user = await db.users.find_one({"email": email})
    if not user:
        new_user = {"email": email, "password_hash": "", "name": name, "role": "operator",
                     "is_active": True, "sso_provider": "google", "created_at": datetime.now(timezone.utc)}
        await db.users.insert_one(new_user)
        user = await db.users.find_one({"email": email})
    
    token = create_token({"sub": str(user.get("id", user.get("email"))), "email": email, "role": user.get("role", "operator")})
    response = Response(status_code=302)
    response.headers["Location"] = "/"
    response.set_cookie(key="session_token", value=token, httponly=True, secure=True, samesite="lax", max_age=86400)
    return response

@api_router.get("/auth/sso/microsoft/url")
async def microsoft_sso_url():
    """Get Microsoft OAuth2 authorization URL"""
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings or not settings.get("sso_microsoft_enabled") or not settings.get("sso_microsoft_client_id"):
        raise HTTPException(status_code=400, detail="SSO Microsoft non configure")
    client_id = settings["sso_microsoft_client_id"]
    tenant_id = settings.get("sso_microsoft_tenant_id", "common")
    redirect_uri = os.environ.get("SSO_REDIRECT_URL", "") + "/api/auth/sso/microsoft/callback"
    scope = "openid email profile User.Read"
    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/authorize?client_id={client_id}&redirect_uri={redirect_uri}&response_type=code&scope={scope}"
    return {"url": url}

@api_router.get("/auth/sso/microsoft/callback")
async def microsoft_sso_callback(code: str, response: Response):
    """Handle Microsoft OAuth2 callback"""
    import httpx
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings or not settings.get("sso_microsoft_enabled"):
        raise HTTPException(status_code=400, detail="SSO Microsoft non active")
    
    client_id = settings["sso_microsoft_client_id"]
    client_secret = settings.get("sso_microsoft_client_secret", "")
    tenant_id = settings.get("sso_microsoft_tenant_id", "common")
    redirect_uri = os.environ.get("SSO_REDIRECT_URL", "") + "/api/auth/sso/microsoft/callback"
    
    async with httpx.AsyncClient() as client:
        token_resp = await client.post(f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token", data={
            "code": code, "client_id": client_id, "client_secret": client_secret,
            "redirect_uri": redirect_uri, "grant_type": "authorization_code", "scope": "openid email profile User.Read"
        })
        if token_resp.status_code != 200:
            raise HTTPException(status_code=400, detail="Erreur d'authentification Microsoft")
        tokens = token_resp.json()
        
        userinfo_resp = await client.get("https://graph.microsoft.com/v1.0/me",
            headers={"Authorization": f"Bearer {tokens['access_token']}"})
        userinfo = userinfo_resp.json()
    
    email = userinfo.get("mail", userinfo.get("userPrincipalName", ""))
    name = userinfo.get("displayName", email)
    
    user = await db.users.find_one({"email": email})
    if not user:
        new_user = {"email": email, "password_hash": "", "name": name, "role": "operator",
                     "is_active": True, "sso_provider": "microsoft", "created_at": datetime.now(timezone.utc)}
        await db.users.insert_one(new_user)
        user = await db.users.find_one({"email": email})
    
    token = create_token({"sub": str(user.get("id", user.get("email"))), "email": email, "role": user.get("role", "operator")})
    response = Response(status_code=302)
    response.headers["Location"] = "/"
    response.set_cookie(key="session_token", value=token, httponly=True, secure=True, samesite="lax", max_age=86400)
    return response

@api_router.get("/auth/sso/status")
async def sso_status():
    """Get SSO status for login page (public)"""
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    if not settings:
        return {"google_enabled": False, "microsoft_enabled": False}
    return {
        "google_enabled": bool(settings.get("sso_google_enabled") and settings.get("sso_google_client_id")),
        "microsoft_enabled": bool(settings.get("sso_microsoft_enabled") and settings.get("sso_microsoft_client_id")),
    }

# ================= PRICE HISTORY =================

@api_router.post("/price-history/record")
async def record_price_history(manager: dict = Depends(require_manager_or_admin)):
    """Record current prices for all recipes"""
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(5000)
    recorded = 0
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe["id"])
            entry = {
                "id": str(uuid.uuid4()),
                "recipe_id": recipe["id"],
                "recipe_name": recipe["name"],
                "supplier_name": recipe.get("supplier_name", ""),
                "version": recipe.get("version", 1),
                "cost_per_unit": round(cost.cost_per_unit, 4),
                "total_cost": round(cost.total_cost, 4),
                "suggested_price": round(cost.suggested_price, 4),
                "recorded_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.price_history.insert_one(entry)
            recorded += 1
        except:
            continue
    return {"recorded": recorded}

@api_router.get("/price-history")
async def get_price_history(recipe_id: Optional[str] = None, days: int = 90):
    """Get price history for dashboard charts"""
    query = {}
    if recipe_id:
        query["recipe_id"] = recipe_id
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    query["recorded_at"] = {"$gte": cutoff}
    entries = await db.price_history.find(query, {"_id": 0}).sort("recorded_at", 1).to_list(10000)
    return entries

@api_router.get("/price-history/alerts")
async def get_price_alerts():
    """Check for price increases above threshold"""
    settings = await db.settings.find_one({"key": "app_settings"}, {"_id": 0})
    threshold = (settings or {}).get("alert_threshold", 10)
    
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(5000)
    alerts = []
    for mat in materials:
        history = await db.price_history_materials.find(
            {"material_id": mat["id"]}, {"_id": 0}
        ).sort("recorded_at", -1).to_list(2)
        if len(history) >= 2:
            old_price = history[1].get("unit_price", 0)
            new_price = history[0].get("unit_price", 0)
            if old_price > 0:
                change_pct = ((new_price - old_price) / old_price) * 100
                if abs(change_pct) >= threshold:
                    alerts.append({
                        "material_name": mat["name"],
                        "supplier_name": mat.get("supplier_name", ""),
                        "old_price": old_price,
                        "new_price": new_price,
                        "change_pct": round(change_pct, 1),
                        "type": "hausse" if change_pct > 0 else "baisse",
                    })
    return alerts

@api_router.post("/price-history/materials/record")
async def record_material_prices(manager: dict = Depends(require_manager_or_admin)):
    """Record current material prices for tracking"""
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(5000)
    recorded = 0
    for mat in materials:
        entry = {
            "id": str(uuid.uuid4()),
            "material_id": mat["id"],
            "material_name": mat["name"],
            "unit_price": mat.get("unit_price", 0),
            "supplier_name": mat.get("supplier_name", ""),
            "recorded_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.price_history_materials.insert_one(entry)
        recorded += 1
    return {"recorded": recorded}

# ================= SIMULATION (WHAT-IF) =================

@api_router.post("/simulation/what-if")
async def simulation_what_if(request: Request):
    """Simulate impact of price change on recipes"""
    data = await request.json()
    material_id = data.get("material_id")
    price_change_pct = data.get("price_change_pct", 0)
    
    if not material_id:
        raise HTTPException(status_code=400, detail="material_id requis")
    
    material = await db.raw_materials.find_one({"id": material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Matiere non trouvee")
    
    original_price = material.get("unit_price", 0)
    new_price = original_price * (1 + price_change_pct / 100)
    
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(5000)
    impacts = []
    
    for recipe in recipes:
        ingredients = recipe.get("ingredients", [])
        uses_material = any(i.get("material_id") == material_id for i in ingredients)
        if not uses_material:
            continue
        
        try:
            current_cost = await calculate_cost(recipe["id"])
            
            # Temporarily change price in memory to calculate new cost
            mat_qty = 0
            for ing in ingredients:
                if ing.get("material_id") == material_id:
                    mat_qty += ing.get("quantity", 0)
            
            price_diff_per_unit = (new_price - original_price) * mat_qty
            freinte = material.get("freinte", 0)
            if freinte > 0:
                price_diff_per_unit = price_diff_per_unit / (1 - freinte / 100)
            
            new_total_cost = current_cost.total_cost + price_diff_per_unit
            output_qty = recipe.get("output_quantity", 1)
            new_cost_per_unit = new_total_cost / output_qty if output_qty > 0 else 0
            margin = recipe.get("target_margin", 30)
            new_suggested_price = new_cost_per_unit / (1 - margin / 100) if margin < 100 else 0
            
            impacts.append({
                "recipe_id": recipe["id"],
                "recipe_name": recipe["name"],
                "supplier_name": recipe.get("supplier_name", ""),
                "version": recipe.get("version", 1),
                "current_cost_per_unit": round(current_cost.cost_per_unit, 2),
                "new_cost_per_unit": round(new_cost_per_unit, 2),
                "cost_diff": round(new_cost_per_unit - current_cost.cost_per_unit, 2),
                "cost_diff_pct": round(((new_cost_per_unit - current_cost.cost_per_unit) / current_cost.cost_per_unit * 100) if current_cost.cost_per_unit > 0 else 0, 1),
                "current_suggested_price": round(current_cost.suggested_price, 2),
                "new_suggested_price": round(new_suggested_price, 2),
            })
        except:
            continue
    
    return {
        "material_name": material["name"],
        "original_price": original_price,
        "new_price": round(new_price, 4),
        "price_change_pct": price_change_pct,
        "impacted_recipes": len(impacts),
        "impacts": impacts,
    }

# ================= SITES MANAGEMENT =================

@api_router.get("/sites")
async def get_sites():
    sites = await db.sites.find({}, {"_id": 0}).to_list(100)
    if not sites:
        default_site = {"id": "default", "name": "Site principal", "address": "", "is_default": True}
        await db.sites.insert_one({**default_site})
        return [default_site]
    return sites

@api_router.post("/sites")
async def create_site(request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    site = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", ""),
        "address": data.get("address", ""),
        "is_default": False,
    }
    await db.sites.insert_one(site)
    return await db.sites.find_one({"id": site["id"]}, {"_id": 0})

@api_router.put("/sites/{site_id}")
async def update_site(site_id: str, request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    data.pop("id", None)
    data.pop("_id", None)
    await db.sites.update_one({"id": site_id}, {"$set": data})
    return await db.sites.find_one({"id": site_id}, {"_id": 0})

@api_router.delete("/sites/{site_id}")
async def delete_site(site_id: str, admin: dict = Depends(require_admin)):
    site = await db.sites.find_one({"id": site_id}, {"_id": 0})
    if site and site.get("is_default"):
        raise HTTPException(status_code=400, detail="Impossible de supprimer le site par defaut")
    await db.sites.delete_one({"id": site_id})
    return {"message": "Site supprime"}

# ================= PUBLIC KPI API =================

async def validate_api_key(request: Request):
    """Validate API key from header or query param"""
    api_key = request.headers.get("X-API-Key") or request.query_params.get("api_key")
    if not api_key:
        raise HTTPException(status_code=401, detail="Cle API requise (header X-API-Key ou param ?api_key=)")
    key_doc = await db.api_keys.find_one({"key": api_key, "is_active": True}, {"_id": 0})
    if not key_doc:
        raise HTTPException(status_code=403, detail="Cle API invalide ou desactivee")
    await db.api_keys.update_one({"key": api_key}, {"$set": {"last_used": datetime.now(timezone.utc).isoformat()}})
    return key_doc

@api_router.get("/public/kpi/doc")
async def kpi_documentation():
    """Documentation complète de l'API KPI publique"""
    return {
        "title": "PrixRevient - API KPI Publique",
        "version": "1.0",
        "authentication": {
            "method": "Cle API",
            "header": "X-API-Key: VOTRE_CLE",
            "query_param": "?api_key=VOTRE_CLE",
            "gestion": "Parametres > API dans l'interface admin"
        },
        "endpoints": [
            {
                "url": "/api/public/kpi/summary",
                "method": "GET",
                "description": "Resume global des KPI",
                "champs_reponse": {
                    "total_recipes": "Nombre total de recettes",
                    "total_materials": "Nombre total de matieres premieres",
                    "total_suppliers": "Nombre total de fournisseurs",
                    "total_categories": "Nombre total de categories",
                    "avg_cost_per_unit": "Cout moyen par unite (EUR)",
                    "avg_margin": "Marge moyenne (%)",
                    "total_production_value": "Valeur totale de production (EUR)",
                    "top_expensive_recipes": "Top 5 recettes les plus couteuses [{name, cost_per_unit, supplier_name, version}]",
                    "top_cheapest_recipes": "Top 5 recettes les moins couteuses [{name, cost_per_unit, supplier_name, version}]",
                    "costs_by_supplier": "Repartition des couts par fournisseur [{supplier, total_cost, recipe_count}]"
                }
            },
            {
                "url": "/api/public/kpi/costs",
                "method": "GET",
                "description": "Detail des couts de toutes les recettes",
                "parametres": {
                    "supplier": "(optionnel) Filtrer par nom de fournisseur",
                    "version": "(optionnel) Filtrer par numero de version"
                },
                "champs_reponse": {
                    "recipe_id": "Identifiant unique de la recette",
                    "recipe_name": "Nom de la recette",
                    "supplier_name": "Nom du fournisseur",
                    "version": "Numero de version",
                    "output_quantity": "Quantite produite",
                    "output_unit": "Unite de mesure",
                    "material_cost": "Cout des matieres (EUR)",
                    "labor_cost": "Cout main d'oeuvre (EUR)",
                    "overhead_cost": "Cout frais generaux (EUR)",
                    "freinte_cost": "Cout de la freinte (EUR)",
                    "total_cost": "Cout total (EUR)",
                    "cost_per_unit": "Cout par unite (EUR)",
                    "target_margin": "Marge cible (%)",
                    "suggested_price": "Prix de vente conseille (EUR)"
                }
            },
            {
                "url": "/api/public/kpi/materials",
                "method": "GET",
                "description": "Liste des matieres premieres avec prix",
                "champs_reponse": {
                    "id": "Identifiant unique",
                    "name": "Nom de la matiere",
                    "unit": "Unite de mesure",
                    "unit_price": "Prix unitaire (EUR)",
                    "supplier_name": "Nom du fournisseur",
                    "category_id": "ID de la categorie",
                    "freinte": "Taux de freinte (%)",
                    "stock_quantity": "Quantite en stock"
                }
            },
            {
                "url": "/api/public/kpi/recipes",
                "method": "GET",
                "description": "Liste des recettes avec details",
                "champs_reponse": {
                    "id": "Identifiant unique",
                    "name": "Nom de la recette",
                    "supplier_name": "Nom du fournisseur",
                    "version": "Numero de version",
                    "output_quantity": "Quantite produite",
                    "output_unit": "Unite de mesure",
                    "target_margin": "Marge cible (%)",
                    "is_intermediate": "Article semi-fini (true/false)",
                    "ingredients_count": "Nombre d'ingredients",
                    "labor_costs_count": "Nombre de postes main d'oeuvre"
                }
            },
            {
                "url": "/api/public/kpi/suppliers",
                "method": "GET",
                "description": "Liste des fournisseurs",
                "champs_reponse": {
                    "id": "Identifiant unique",
                    "name": "Nom du fournisseur",
                    "contact": "Personne de contact",
                    "email": "Email",
                    "phone": "Telephone",
                    "address": "Adresse"
                }
            }
        ],
        "exemple_curl": "curl -H 'X-API-Key: VOTRE_CLE' https://calculprix.appli-sciad.com/api/public/kpi/summary"
    }

@api_router.get("/public/kpi/summary")
async def kpi_summary(api_key_data: dict = Depends(validate_api_key)):
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(5000)
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(5000)
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    categories = await db.categories.find({}, {"_id": 0}).to_list(1000)

    costs_data = []
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe["id"])
            costs_data.append({
                "name": recipe["name"],
                "supplier_name": recipe.get("supplier_name", ""),
                "version": recipe.get("version", 1),
                "cost_per_unit": cost.cost_per_unit,
                "total_cost": cost.total_cost,
                "target_margin": cost.target_margin,
            })
        except:
            continue

    avg_cost = sum(c["cost_per_unit"] for c in costs_data) / len(costs_data) if costs_data else 0
    avg_margin = sum(c["target_margin"] for c in costs_data) / len(costs_data) if costs_data else 0
    total_value = sum(c["total_cost"] for c in costs_data)

    sorted_by_cost = sorted(costs_data, key=lambda x: x["cost_per_unit"], reverse=True)
    by_supplier = {}
    for c in costs_data:
        s = c["supplier_name"] or "Sans fournisseur"
        if s not in by_supplier:
            by_supplier[s] = {"supplier": s, "total_cost": 0, "recipe_count": 0}
        by_supplier[s]["total_cost"] += c["total_cost"]
        by_supplier[s]["recipe_count"] += 1

    return {
        "total_recipes": len(recipes),
        "total_materials": len(materials),
        "total_suppliers": len(suppliers),
        "total_categories": len(categories),
        "avg_cost_per_unit": round(avg_cost, 2),
        "avg_margin": round(avg_margin, 1),
        "total_production_value": round(total_value, 2),
        "top_expensive_recipes": sorted_by_cost[:5],
        "top_cheapest_recipes": sorted_by_cost[-5:][::-1] if len(sorted_by_cost) >= 5 else sorted_by_cost[::-1],
        "costs_by_supplier": list(by_supplier.values()),
    }

@api_router.get("/public/kpi/costs")
async def kpi_costs(supplier: Optional[str] = None, version: Optional[int] = None, api_key_data: dict = Depends(validate_api_key)):
    query = {}
    if supplier:
        query["supplier_name"] = supplier
    if version:
        query["version"] = version
    recipes = await db.recipes.find(query, {"_id": 0}).to_list(5000)
    results = []
    for recipe in recipes:
        try:
            cost = await calculate_cost(recipe["id"])
            results.append({
                "recipe_id": recipe["id"],
                "recipe_name": recipe["name"],
                "supplier_name": recipe.get("supplier_name", ""),
                "version": recipe.get("version", 1),
                "output_quantity": recipe.get("output_quantity", 1),
                "output_unit": recipe.get("output_unit", "piece"),
                "material_cost": round(cost.total_material_cost, 2),
                "labor_cost": round(cost.total_labor_cost, 2),
                "overhead_cost": round(cost.total_overhead_cost, 2),
                "freinte_cost": round(cost.total_freinte_cost, 2),
                "total_cost": round(cost.total_cost, 2),
                "cost_per_unit": round(cost.cost_per_unit, 2),
                "target_margin": cost.target_margin,
                "suggested_price": round(cost.suggested_price, 2),
            })
        except:
            continue
    return results

@api_router.get("/public/kpi/materials")
async def kpi_materials(api_key_data: dict = Depends(validate_api_key)):
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(5000)
    return [{"id": m["id"], "name": m["name"], "unit": m.get("unit",""), "unit_price": m.get("unit_price",0),
             "supplier_name": m.get("supplier_name",""), "category_id": m.get("category_id",""),
             "freinte": m.get("freinte",0), "stock_quantity": m.get("stock_quantity",0)} for m in materials]

@api_router.get("/public/kpi/recipes")
async def kpi_recipes(api_key_data: dict = Depends(validate_api_key)):
    recipes = await db.recipes.find({}, {"_id": 0}).to_list(5000)
    return [{"id": r["id"], "name": r["name"], "supplier_name": r.get("supplier_name",""),
             "version": r.get("version",1), "output_quantity": r.get("output_quantity",1),
             "output_unit": r.get("output_unit","piece"), "target_margin": r.get("target_margin",30),
             "is_intermediate": r.get("is_intermediate",False),
             "ingredients_count": len(r.get("ingredients",[])), "labor_costs_count": len(r.get("labor_costs",[]))} for r in recipes]

@api_router.get("/public/kpi/suppliers")
async def kpi_suppliers(api_key_data: dict = Depends(validate_api_key)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    return [{"id": s["id"], "name": s["name"], "contact": s.get("contact",""),
             "email": s.get("email",""), "phone": s.get("phone",""), "address": s.get("address","")} for s in suppliers]

# ================= API KEYS MANAGEMENT =================

@api_router.get("/api-keys")
async def get_api_keys(admin: dict = Depends(require_admin)):
    keys = await db.api_keys.find({}, {"_id": 0}).to_list(100)
    return keys

@api_router.post("/api-keys")
async def create_api_key(request: Request, admin: dict = Depends(require_admin)):
    data = await request.json()
    key = {
        "id": str(uuid.uuid4()),
        "name": data.get("name", "Cle API"),
        "key": f"pk_{uuid.uuid4().hex}",
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_used": None,
    }
    await db.api_keys.insert_one(key)
    return await db.api_keys.find_one({"id": key["id"]}, {"_id": 0})

@api_router.delete("/api-keys/{key_id}")
async def delete_api_key(key_id: str, admin: dict = Depends(require_admin)):
    await db.api_keys.delete_one({"id": key_id})
    return {"message": "Cle API supprimee"}

@api_router.put("/api-keys/{key_id}/toggle")
async def toggle_api_key(key_id: str, admin: dict = Depends(require_admin)):
    key_doc = await db.api_keys.find_one({"id": key_id}, {"_id": 0})
    if not key_doc:
        raise HTTPException(status_code=404, detail="Cle non trouvee")
    new_status = not key_doc.get("is_active", True)
    await db.api_keys.update_one({"id": key_id}, {"$set": {"is_active": new_status}})
    return {"is_active": new_status}

# ================= APP SETUP =================
app.include_router(api_router)

frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[frontend_url, "http://localhost:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    await db.users.create_index("email", unique=True)
    
    # Only seed admin if NO admin user exists at all
    any_admin = await db.users.find_one({"role": "admin"})
    if not any_admin:
        admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com")
        admin_password = os.environ.get("ADMIN_PASSWORD", "Admin123!")
        existing = await db.users.find_one({"email": admin_email})
        if not existing:
            hashed = hash_password(admin_password)
            await db.users.insert_one({"email": admin_email, "password_hash": hashed, "name": "Admin", "role": "admin", "is_active": True, "created_at": datetime.now(timezone.utc)})
            logger.info(f"Admin user created: {admin_email}")
    
    # Start background scheduler
    await sync_scheduler()

@app.on_event("shutdown")
async def shutdown_db_client():
    if _scheduler_started:
        scheduler.shutdown(wait=False)
    client.close()
